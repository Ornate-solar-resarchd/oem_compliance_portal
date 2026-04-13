"""
OEM Datasheet Extraction Engine
Extracts technical specs from uploaded datasheets (PDF/Excel) using keyword pattern matching.
Supports categories: Cell, DC Block, PCS, EMS — no AI dependency.

PDF Parsing: PyMuPDF (fitz) for text + Camelot for table extraction.
This gives 2-3x better extraction from BESS datasheets which are mostly tables.
"""
import os
import json
import re
import io

CATEGORY_PROMPTS = {
    "Cell": """Extract ALL battery cell technical specifications:
- Nominal Capacity (Ah), Nominal Voltage (V), Energy (Wh/kWh)
- Internal Resistance (mΩ), Cycle Life, Calendar Life
- Max Charge/Discharge Rate (C-rate), Max Current (A)
- Weight (kg), Dimensions (L×W×H mm), Energy Density (Wh/kg)
- Operating Temperature range (charge/discharge/storage)
- Chemistry (LFP/NMC/etc), Form Factor (Prismatic/Cylindrical/Pouch/Blade)
- Certifications (IEC 62619, UL 1973, UN 38.3, BIS, UL 9540A)
- Self-discharge rate, Coulombic efficiency, Round-trip efficiency
- EOL capacity retention, SOC operating range
- Any other specs mentioned""",

    "DC Block": """Extract ALL DC Block / Battery Container / DC System specifications:
- Rated Energy (kWh/MWh), Rated Power (kW/MW)
- DC Voltage range (min/max/nominal)
- Number of battery modules/racks/clusters
- Cooling type (liquid/air), Cooling capacity
- Container dimensions (L×W×H), Weight
- Protection rating (IP rating), Operating temperature
- BMS type, Communication protocol
- Fire suppression system type
- Certifications, Safety features
- Any other specs mentioned""",

    "PCS": """Extract ALL Power Conversion System (PCS) / Inverter specifications:
- Rated AC Power (kW/MW), Rated DC Power
- AC Voltage (V), DC Voltage range
- Frequency (Hz), Power Factor range
- Max Efficiency, Euro/CEC efficiency
- THD (Total Harmonic Distortion)
- Cooling type, Operating temperature
- Dimensions, Weight
- Communication protocols (Modbus/CAN/Ethernet)
- Grid support features (PFR, SFR, Black Start, Reactive Power)
- Certifications (IEC 62477, IEC 61000, IEEE 1547)
- Protection features
- Any other specs mentioned""",

    "EMS": """Extract ALL Energy Management System (EMS) / SCADA specifications:
- System architecture (server config, redundancy)
- Data refresh rate, System availability (%)
- Communication protocols (IEC 61850, Modbus TCP/IP, DNP3)
- Control modes (peak shaving, frequency regulation, load shifting, etc.)
- SCADA integration, HMI details
- Cyber security features (RBAC, firewall, audit trail)
- GPS time synchronization
- Remote monitoring capabilities
- Software origin (indigenous/imported)
- Hardware requirements (servers, RTUs, gateways)
- Certifications (CEA, IEC standards)
- Any other specs mentioned""",
}


def _normalize_text(text: str) -> str:
    """Normalize extracted PDF text for regex matching.
    PyMuPDF preserves line breaks from PDF layout, which breaks .*? regex patterns.
    We collapse mid-sentence newlines to spaces while keeping paragraph breaks."""
    # Replace single newlines (mid-sentence breaks) with spaces
    # Keep double newlines (paragraph breaks)
    text = re.sub(r'(?<!\n)\n(?!\n)', ' ', text)
    # Collapse multiple spaces
    text = re.sub(r'  +', ' ', text)
    return text


def _extract_pdf_pymupdf(contents: bytes) -> str:
    """Extract text from PDF using PyMuPDF (fitz) — much better layout preservation than PyPDF2."""
    try:
        import fitz
        doc = fitz.open(stream=contents, filetype="pdf")
        pages = []
        for page in doc:
            pages.append(page.get_text("text"))
        doc.close()
        raw = "\n".join(pages)
        return _normalize_text(raw)
    except Exception as e:
        print(f"[PyMuPDF] extraction error: {e}")
        return ""


def _extract_pdf_tables_camelot(contents: bytes) -> str:
    """Extract tables from PDF using Camelot — converts tabular data into clean key:value text.
    This is critical for BESS datasheets which are 80%+ tables."""
    import tempfile
    tmp_path = None
    try:
        import camelot
        # Camelot needs a file path, not bytes
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
            tmp.write(contents)
            tmp_path = tmp.name

        # Try lattice (grid-based tables) first, then stream (text-based)
        tables = camelot.read_pdf(tmp_path, pages="all", flavor="lattice")
        if len(tables) == 0:
            tables = camelot.read_pdf(tmp_path, pages="all", flavor="stream")

        if len(tables) == 0:
            return ""

        print(f"[Camelot] Extracted {len(tables)} tables from PDF")

        table_lines = []
        for table in tables:
            df = table.df
            # Convert table rows to "key: value" or "key value unit" format
            for _, row in df.iterrows():
                cells = [str(c).strip() for c in row if str(c).strip()]
                if len(cells) >= 2:
                    # Join cells with separator — makes regex matching work on table data
                    table_lines.append(" ".join(cells))
                elif len(cells) == 1:
                    table_lines.append(cells[0])

        return "\n".join(table_lines)

    except Exception as e:
        print(f"[Camelot] table extraction error: {e}")
        return ""
    finally:
        if tmp_path:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass


def extract_text_from_file(contents: bytes, filename: str) -> str:
    """Extract text from PDF or Excel file.
    For PDFs: Uses PyMuPDF for text + Camelot for tables, merged together.
    This gives much better results on BESS datasheets than PyPDF2 alone."""
    ext = (filename or "").rsplit(".", 1)[-1].lower()

    if ext == "pdf":
        # Primary: PyMuPDF for full text
        text = _extract_pdf_pymupdf(contents)

        # Secondary: Camelot for tables (merged with main text)
        table_text = _extract_pdf_tables_camelot(contents)

        if table_text:
            # Merge: main text first, then table-extracted text
            # This ensures regex patterns match both freeform text AND tabular data
            combined = text + "\n\n--- EXTRACTED TABLES ---\n\n" + table_text
            print(f"[PDF] PyMuPDF: {len(text)} chars + Camelot tables: {len(table_text)} chars = {len(combined)} chars total")
            return combined

        # Fallback to PyPDF2 if PyMuPDF fails
        if not text:
            try:
                from PyPDF2 import PdfReader
                reader = PdfReader(io.BytesIO(contents))
                text = "\n".join(p.extract_text() or "" for p in reader.pages)
                print(f"[PDF] Fallback PyPDF2: {len(text)} chars")
            except Exception as e:
                print(f"[PDF] All extraction methods failed: {e}")

        return text

    elif ext in ("xlsx", "xls"):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(contents), data_only=True)
            lines = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    row_text = " ".join(str(c) for c in row if c is not None)
                    if row_text.strip():
                        lines.append(row_text)
            return "\n".join(lines)
        except Exception as e:
            print(f"Excel extraction error: {e}")
            return ""

    elif ext == "csv":
        return contents.decode("utf-8", errors="ignore")

    return ""


def extract_specs_with_gemini(text: str, category: str, api_key: str) -> list:
    """Use Gemini AI to extract specs from datasheet text."""
    try:
        import google.generativeai as genai
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel("gemini-2.0-flash")

        category_prompt = CATEGORY_PROMPTS.get(category, CATEGORY_PROMPTS["Cell"])

        prompt = f"""You are an expert BESS (Battery Energy Storage System) technical engineer.
Your job: Extract EVERY SINGLE technical specification, parameter, value, rating, and measurement from this {category} datasheet.

BE EXHAUSTIVE. Extract EVERYTHING — even minor specs. The goal is to capture 30-50+ parameters minimum.

{category_prompt}

IMPORTANT RULES:
1. Extract EVERY number, rating, spec, measurement, range, limit, threshold mentioned in the document
2. If a parameter has a range (e.g., "0°C to 55°C"), extract both min and max as separate entries
3. Include ALL certifications, standards, and compliance marks individually
4. Include model number, manufacturer, part number, dimensions, weight
5. Include operating conditions, storage conditions, transportation conditions
6. Include warranty, design life, degradation specs
7. Include communication protocols, connector types, mounting details
8. If something is mentioned but no exact value, extract it with "See datasheet" as value
9. DO NOT skip any specification — more is better

DATASHEET TEXT:
{text[:40000]}

Return a JSON array of objects, each with:
- "name": parameter name (human readable, descriptive)
- "code": short uppercase code (e.g. CELL_CAPACITY_AH, PCS_RATED_POWER, DC_COOLING_TYPE)
- "value": the exact extracted value as string
- "unit": unit of measurement (V, A, Ah, mm, kg, °C, %, W, kW, MW, MWh, etc.)
- "section": category (Electrical, Physical, Thermal, Safety, Performance, Communication, Mechanical, General)
- "status": "pass"

CRITICAL: Return AT LEAST 20 parameters. Extract every single data point. Return ONLY the JSON array. Start with [ end with ]."""

        response = model.generate_content(prompt)
        response_text = response.text.strip()

        if response_text.startswith("```"):
            response_text = re.sub(r'^```(?:json)?\s*', '', response_text)
            response_text = re.sub(r'\s*```$', '', response_text)

        if response_text.startswith("["):
            return json.loads(response_text)

        match = re.search(r'\[.*\]', response_text, re.DOTALL)
        if match:
            return json.loads(match.group())

    except Exception as e:
        print(f"Gemini datasheet extraction error: {e}")

    return []


def extract_specs_with_claude(text: str, category: str, api_key: str) -> list:
    """Use Claude API to extract specs from datasheet text (primary)."""
    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)

        category_prompt = CATEGORY_PROMPTS.get(category, CATEGORY_PROMPTS["Cell"])

        prompt = f"""You are an expert BESS technical engineer. Extract EVERY technical specification from this {category} datasheet.

BE EXHAUSTIVE — extract 30-50+ parameters minimum.

{category_prompt}

RULES:
1. Extract EVERY number, rating, spec, measurement, range, limit, threshold
2. Ranges → extract min and max as separate entries
3. Include ALL certifications individually
4. Include model number, manufacturer, dimensions, weight
5. Include warranty, design life, degradation specs
6. DO NOT skip any specification

DATASHEET TEXT:
{text[:40000]}

Return a JSON array. Each object must have:
- "name": parameter name (human readable)
- "code": short uppercase code (e.g. CELL_CAPACITY_AH)
- "value": exact extracted value as string
- "unit": unit of measurement
- "section": category (Electrical, Physical, Thermal, Safety, Performance, Communication, General)
- "status": "pass"
- "page": approximate page number where found (number or null)
- "source_text": exact phrase from document where value was found (max 80 chars)

Return AT LEAST 20 parameters. Return ONLY the JSON array."""

        message = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=4096,
            messages=[{"role": "user", "content": prompt}],
        )

        response_text = message.content[0].text.strip()
        if response_text.startswith("```"):
            response_text = re.sub(r'^```(?:json)?\s*', '', response_text)
            response_text = re.sub(r'\s*```$', '', response_text)
        if response_text.startswith("["):
            return json.loads(response_text)
        match = re.search(r'\[.*\]', response_text, re.DOTALL)
        if match:
            return json.loads(match.group())

    except Exception as e:
        print(f"Claude datasheet extraction error: {e}")

    return []


def extract_specs_keyword(text: str, category: str) -> list:
    """Comprehensive keyword-based extraction — 50-70+ patterns per category."""
    text_lower = text.lower()
    extracted = []

    if category == "Cell":
        patterns = [
            # ── Electrical ──
            ("Nominal Capacity", "CELL_CAPACITY_AH", r"(?:nominal|rated|typical)\s*capacity.*?(\d+(?:\.\d+)?)\s*Ah", "Ah", "Electrical"),
            ("Minimum Capacity", "CELL_MIN_CAPACITY_AH", r"(?:minimum|min)\s*capacity.*?(\d+(?:\.\d+)?)\s*Ah", "Ah", "Electrical"),
            ("Nominal Voltage", "CELL_VOLTAGE_V", r"(?:nominal|rated)\s*voltage.*?(\d+(?:\.\d+)?)\s*V", "V", "Electrical"),
            ("Charge Cut-off Voltage", "CELL_CHARGE_CUTOFF_V", r"(?:charge|charging)\s*(?:cut[\s-]*off|upper)\s*voltage.*?(\d+(?:\.\d+)?)\s*V", "V", "Electrical"),
            ("Discharge Cut-off Voltage", "CELL_DISCHARGE_CUTOFF_V", r"(?:discharge|discharging)\s*(?:cut[\s-]*off|lower)\s*voltage.*?(\d+(?:\.\d+)?)\s*V", "V", "Electrical"),
            ("Energy per Cell", "CELL_ENERGY_WH", r"energy.*?(\d+(?:\.\d+)?)\s*(?:Wh|kWh)", "Wh", "Electrical"),
            ("Internal Resistance (AC)", "CELL_IR_AC_MOHM", r"(?:AC\s*)?(?:internal\s*)?resistance.*?(\d+(?:\.\d+)?)\s*m[Ωo]", "mΩ", "Electrical"),
            ("Internal Resistance (DC)", "CELL_IR_DC_MOHM", r"DC\s*(?:internal\s*)?resistance.*?(\d+(?:\.\d+)?)\s*m[Ωo]", "mΩ", "Electrical"),
            ("Max Charge Current", "CELL_MAX_CHARGE_A", r"(?:max|maximum)\s*(?:charge|charging)\s*current.*?(\d+(?:\.\d+)?)\s*A", "A", "Electrical"),
            ("Max Discharge Current", "CELL_MAX_DISCHARGE_A", r"(?:max|maximum)\s*(?:discharge|discharging)\s*current.*?(\d+(?:\.\d+)?)\s*A", "A", "Electrical"),
            ("Standard Charge Current", "CELL_STD_CHARGE_A", r"(?:standard|recommended|normal)\s*(?:charge|charging)\s*current.*?(\d+(?:\.\d+)?)\s*A", "A", "Electrical"),
            ("Standard Discharge Current", "CELL_STD_DISCHARGE_A", r"(?:standard|recommended|normal)\s*(?:discharge|discharging)\s*current.*?(\d+(?:\.\d+)?)\s*A", "A", "Electrical"),
            ("Max Charge Rate", "CELL_MAX_CHARGE_RATE", r"(?:max|maximum)\s*(?:charge|charging)\s*rate.*?(\d+(?:\.\d+)?)\s*C", "C", "Electrical"),
            ("Max Discharge Rate", "CELL_MAX_DISCHARGE_RATE", r"(?:max|maximum)\s*(?:discharge|discharging)\s*rate.*?(\d+(?:\.\d+)?)\s*C", "C", "Electrical"),
            ("Short Circuit Current", "CELL_SHORT_CIRCUIT_A", r"short[\s-]*circuit\s*current.*?(\d+(?:\.\d+)?)\s*A", "A", "Electrical"),
            ("Open Circuit Voltage", "CELL_OCV_V", r"open[\s-]*circuit\s*voltage.*?(\d+(?:\.\d+)?)\s*V", "V", "Electrical"),
            ("SOC Range", "CELL_SOC_RANGE", r"SOC.*?(\d+)\s*%?\s*[-~to]+\s*(\d+)\s*%", "%", "Electrical"),
            # ── Performance / Life ──
            ("Cycle Life", "CELL_CYCLE_LIFE", r"cycle\s*life.*?(\d[\d,]*)\s*(?:cycles|times)", "cycles", "Performance"),
            ("Calendar Life", "CELL_CALENDAR_LIFE", r"calendar\s*life.*?(\d+)\s*years?", "years", "Performance"),
            ("EOL Capacity Retention", "CELL_EOL_RETENTION", r"(?:EOL|end\s*of\s*life|capacity\s*retention).*?(\d+(?:\.\d+)?)\s*%", "%", "Performance"),
            ("Self-Discharge Rate", "CELL_SELF_DISCHARGE", r"self[\s-]*discharge.*?(\d+(?:\.\d+)?)\s*%", "%/month", "Performance"),
            ("Coulombic Efficiency", "CELL_COULOMBIC_EFF", r"coulombic\s*efficiency.*?(\d+(?:\.\d+)?)\s*%", "%", "Performance"),
            ("Round-trip Efficiency", "CELL_RTE", r"(?:round[\s-]*trip|energy)\s*efficiency.*?(\d+(?:\.\d+)?)\s*%", "%", "Performance"),
            ("Annual Degradation Rate", "CELL_DEGRADATION", r"(?:annual\s*)?degradation.*?(\d+(?:\.\d+)?)\s*%", "%/year", "Performance"),
            ("DOD (Depth of Discharge)", "CELL_DOD", r"(?:DOD|depth\s*of\s*discharge).*?(\d+(?:\.\d+)?)\s*%", "%", "Performance"),
            ("Capacity Fade per 1000 Cycles", "CELL_FADE_RATE", r"(?:capacity\s*)?fade.*?(\d+(?:\.\d+)?)\s*%.*?(?:1000|per)", "%", "Performance"),
            # ── Physical ──
            ("Weight", "CELL_WEIGHT_KG", r"weight.*?(\d+(?:\.\d+)?)\s*(?:kg|g)", "kg", "Physical"),
            ("Length", "CELL_LENGTH_MM", r"length.*?(\d+(?:\.\d+)?)\s*mm", "mm", "Physical"),
            ("Width", "CELL_WIDTH_MM", r"width.*?(\d+(?:\.\d+)?)\s*mm", "mm", "Physical"),
            ("Height / Thickness", "CELL_HEIGHT_MM", r"(?:height|thickness).*?(\d+(?:\.\d+)?)\s*mm", "mm", "Physical"),
            ("Energy Density (Gravimetric)", "CELL_ENERGY_DENSITY_WH_KG", r"(?:energy|gravimetric)\s*density.*?(\d+(?:\.\d+)?)\s*Wh/kg", "Wh/kg", "Physical"),
            ("Energy Density (Volumetric)", "CELL_ENERGY_DENSITY_WH_L", r"(?:volumetric)\s*.*?density.*?(\d+(?:\.\d+)?)\s*Wh/[Ll]", "Wh/L", "Physical"),
            ("Terminal Type", "CELL_TERMINAL", r"terminal.*?(M\d+|bolt|screw|threaded)", "", "Physical"),
            # ── Thermal ──
            ("Charge Temp Min", "CELL_CHARGE_TEMP_MIN", r"(?:charge|charging)\s*(?:temp|temperature).*?(-?\d+)\s*[°℃C]", "°C", "Thermal"),
            ("Charge Temp Max", "CELL_CHARGE_TEMP_MAX", r"(?:charge|charging)\s*(?:temp|temperature).*?(\d+)\s*[°℃C]", "°C", "Thermal"),
            ("Discharge Temp Min", "CELL_DISCHARGE_TEMP_MIN", r"(?:discharge|discharging)\s*(?:temp|temperature).*?(-?\d+)\s*[°℃C]", "°C", "Thermal"),
            ("Discharge Temp Max", "CELL_DISCHARGE_TEMP_MAX", r"(?:discharge|discharging)\s*(?:temp|temperature).*?(\d+)\s*[°℃C]", "°C", "Thermal"),
            ("Storage Temp Min", "CELL_STORAGE_TEMP_MIN", r"storage\s*(?:temp|temperature).*?(-?\d+)\s*[°℃C]", "°C", "Thermal"),
            ("Storage Temp Max", "CELL_STORAGE_TEMP_MAX", r"storage\s*(?:temp|temperature).*?(\d+)\s*[°℃C]", "°C", "Thermal"),
            ("Thermal Runaway Onset", "CELL_TR_ONSET", r"(?:thermal\s*runaway|TR)\s*(?:onset)?.*?(\d+(?:\.\d+)?)\s*[°℃C]", "°C", "Safety"),
            ("Venting Temperature", "CELL_VENT_TEMP", r"vent(?:ing)?\s*(?:onset|temp).*?(\d+(?:\.\d+)?)\s*[°℃C]", "°C", "Safety"),
            # ── Module / Pack ──
            ("Cells per Module", "CELL_PER_MODULE", r"(\d+)\s*cells?\s*per\s*module", "", "System"),
            ("Modules per Rack", "CELL_MOD_PER_RACK", r"(\d+)\s*modules?\s*per\s*rack", "", "System"),
            ("Module Voltage", "CELL_MODULE_VOLTAGE", r"module\s*(?:nominal\s*)?voltage.*?(\d+(?:\.\d+)?)\s*V", "V", "System"),
            ("Module Energy", "CELL_MODULE_ENERGY", r"module\s*(?:nominal\s*)?energy.*?(\d+(?:\.\d+)?)\s*(?:kWh|Wh)", "kWh", "System"),
            ("String Configuration", "CELL_STRING_CONFIG", r"(\d+[SP]\d*[SP]?\d*)", "", "System"),
            ("Total Cells", "CELL_TOTAL_COUNT", r"total\s*(?:number\s*of\s*)?cells.*?(\d[\d,]*)", "", "System"),
        ]
    elif category == "DC Block":
        patterns = [
            # ── System ──
            ("Rated Energy", "DC_RATED_ENERGY", r"(?:rated|nominal)\s*energy.*?(\d+(?:\.\d+)?)\s*(?:kWh|MWh)", "kWh", "System"),
            ("Rated Power", "DC_RATED_POWER", r"(?:rated|nominal)\s*power.*?(\d+(?:\.\d+)?)\s*(?:kW|MW)", "kW", "System"),
            ("Usable Energy", "DC_USABLE_ENERGY", r"usable\s*energy.*?(\d+(?:\.\d+)?)\s*(?:kWh|MWh)", "kWh", "System"),
            ("Nameplate Capacity", "DC_NAMEPLATE", r"nameplate\s*(?:capacity|energy).*?(\d+(?:\.\d+)?)\s*(?:kWh|MWh)", "kWh", "System"),
            ("Number of Modules", "DC_MODULE_COUNT", r"(\d+)\s*(?:modules|racks|battery\s*packs)", "", "System"),
            ("Number of Clusters", "DC_CLUSTER_COUNT", r"(\d+)\s*clusters", "", "System"),
            ("Number of Strings", "DC_STRING_COUNT", r"(\d+)\s*(?:battery\s*)?strings", "", "System"),
            ("Cells per Container", "DC_CELLS_TOTAL", r"(?:total\s*)?(\d[\d,]*)\s*cells?\s*(?:per|in)", "", "System"),
            ("Discharge Duration", "DC_DISCHARGE_HRS", r"(?:discharge|storage)\s*(?:duration|time).*?(\d+(?:\.\d+)?)\s*(?:hour|hr|h)", "hours", "System"),
            ("Container Type", "DC_CONTAINER_TYPE", r"(20\s*ft|40\s*ft|ISO)\s*(?:high[\s-]*cube|container|standard)?", "", "System"),
            # ── Electrical ──
            ("DC Voltage Nominal", "DC_VOLTAGE_NOM", r"(?:dc|battery|nominal|system)\s*voltage.*?(\d+(?:\.\d+)?)\s*V", "V", "Electrical"),
            ("DC Voltage Range Max", "DC_VOLTAGE_MAX", r"(?:max|maximum)\s*(?:dc\s*)?voltage.*?(\d+(?:\.\d+)?)\s*V", "V", "Electrical"),
            ("DC Voltage Range Min", "DC_VOLTAGE_MIN", r"(?:min|minimum)\s*(?:dc\s*)?voltage.*?(\d+(?:\.\d+)?)\s*V", "V", "Electrical"),
            ("Max DC Current", "DC_MAX_CURRENT", r"(?:max|maximum)\s*(?:dc\s*)?current.*?(\d+(?:\.\d+)?)\s*A", "A", "Electrical"),
            ("Auxiliary Power Supply", "DC_AUX_POWER", r"auxiliary.*?(?:power|supply).*?(\d+(?:\.\d+)?)\s*(?:kVA|kW|V)", "kVA", "Electrical"),
            ("UPS Backup", "DC_UPS", r"UPS.*?(\d+(?:\.\d+)?)\s*(?:kVA|kW)", "kVA", "Electrical"),
            # ── Performance ──
            ("RTE (DC)", "DC_RTE", r"(?:round[\s-]*trip|RTE|DC[\s-]*RTE).*?(\d+(?:\.\d+)?)\s*%", "%", "Performance"),
            ("Cycle Life", "DC_CYCLE_LIFE", r"cycle\s*life.*?(\d[\d,]*)\s*(?:cycles|times)", "cycles", "Performance"),
            ("Calendar Life", "DC_CALENDAR_LIFE", r"calendar\s*life.*?(\d+)\s*years?", "years", "Performance"),
            ("DOD", "DC_DOD", r"(?:DOD|depth\s*of\s*discharge).*?(\d+(?:\.\d+)?)\s*%", "%", "Performance"),
            ("Annual Degradation", "DC_DEGRADATION", r"(?:annual\s*)?degradation.*?(\d+(?:\.\d+)?)\s*%", "%/year", "Performance"),
            ("SOC Accuracy", "DC_SOC_ACC", r"SOC\s*(?:accuracy|error).*?[±]?\s*(\d+(?:\.\d+)?)\s*%", "%", "Performance"),
            ("Self-Discharge Rate", "DC_SELF_DISCHARGE", r"self[\s-]*discharge.*?(\d+(?:\.\d+)?)\s*%", "%/month", "Performance"),
            # ── Thermal ──
            ("Cooling Type", "DC_COOLING", r"(liquid|air|hybrid)\s*cool", "", "Thermal"),
            ("Cooling Capacity", "DC_COOLING_KW", r"cool.*?(?:capacity|power).*?(\d+(?:\.\d+)?)\s*kW", "kW", "Thermal"),
            ("Coolant Type", "DC_COOLANT", r"(?:coolant|glycol).*?(ethylene\s*glycol|propylene|water[\s-]*glycol)", "", "Thermal"),
            ("Coolant Volume", "DC_COOLANT_VOL", r"(?:coolant|glycol)\s*(?:volume|quantity).*?(\d+)\s*[Ll]", "L", "Thermal"),
            ("HVAC Power", "DC_HVAC_KW", r"HVAC.*?(\d+(?:\.\d+)?)\s*kW", "kW", "Thermal"),
            ("Operating Temp Min", "DC_OP_TEMP_MIN", r"operating\s*temp.*?(-?\d+)\s*[°℃C]", "°C", "Thermal"),
            ("Operating Temp Max", "DC_OP_TEMP_MAX", r"operating\s*temp.*?(\d+)\s*[°℃C]", "°C", "Thermal"),
            ("Charge Temp Range", "DC_CHARGE_TEMP", r"(?:charge|charging)\s*temp.*?(-?\d+)\s*[°℃C]?\s*(?:to|~|–|-)\s*(\d+)\s*[°℃C]", "°C", "Thermal"),
            ("Discharge Temp Range", "DC_DISCHARGE_TEMP", r"(?:discharge|discharging)\s*temp.*?(-?\d+)\s*[°℃C]?\s*(?:to|~|–|-)\s*(\d+)\s*[°℃C]", "°C", "Thermal"),
            ("Altitude Rating", "DC_ALTITUDE", r"(?:altitude|elevation).*?(\d[\d,]*)\s*m", "m", "Thermal"),
            ("Humidity Range", "DC_HUMIDITY", r"humidity.*?(\d+)\s*%?\s*(?:to|~|–|-)\s*(\d+)\s*%", "%RH", "Thermal"),
            # ── Physical ──
            ("Container Dimensions", "DC_DIMENSIONS", r"(\d{3,5})\s*[×xX]\s*(\d{3,5})\s*[×xX]\s*(\d{3,5})\s*mm", "mm", "Physical"),
            ("Container Length", "DC_LENGTH", r"(?:container\s*)?length.*?(\d+(?:\.\d+)?)\s*(?:mm|m|ft)", "mm", "Physical"),
            ("Container Width", "DC_WIDTH", r"(?:container\s*)?width.*?(\d+(?:\.\d+)?)\s*(?:mm|m|ft)", "mm", "Physical"),
            ("Container Height", "DC_HEIGHT", r"(?:container\s*)?height.*?(\d+(?:\.\d+)?)\s*(?:mm|m|ft)", "mm", "Physical"),
            ("Total Weight", "DC_WEIGHT", r"(?:total\s*)?weight.*?(\d+(?:\.\d+)?)\s*(?:kg|ton|t\b)", "kg", "Physical"),
            ("Protection Rating", "DC_IP_RATING", r"(IP\d+\w?)", "", "Safety"),
            # ── Safety ──
            ("Fire Suppression Type", "DC_FIRE_SYSTEM", r"(aerosol|gas|water\s*mist|Novec\s*1230|FM[\s-]*200|dry\s*pipe|sprinkler)", "", "Safety"),
            ("Fire Detection", "DC_FIRE_DETECT", r"(VESDA|smoke\s*detect|heat\s*detect|gas\s*detect)", "", "Safety"),
            ("Explosion-Proof Vent", "DC_EXPLOSION_VENT", r"(explosion[\s-]*proof|deflagration|pressure\s*relief)", "", "Safety"),
            ("Gas Sensors", "DC_GAS_SENSORS", r"(H[₂2]\s*(?:sensor|detect)|CO\s*(?:sensor|detect)|VOC)", "", "Safety"),
            ("Emergency Ventilation", "DC_EMER_VENT", r"emergency\s*ventil.*?(\d+)\s*(?:CFM|m³)", "CFM", "Safety"),
            # ── BMS ──
            ("BMS Architecture", "DC_BMS_ARCH", r"(?:BMS|battery\s*management).*?(3[\s-]*tier|2[\s-]*tier|hierarchi)", "", "Communication"),
            ("BMS Communication", "DC_BMS_COMM", r"(?:BMS|battery\s*management).*?(CAN|RS485|Modbus|Ethernet|IEC\s*61850)", "", "Communication"),
            ("SOH Monitoring", "DC_SOH", r"(SOH|state\s*of\s*health)\s*(?:monitor|track|report)", "", "Communication"),
            ("Cell Balancing", "DC_CELL_BALANCE", r"(active|passive)\s*(?:cell\s*)?balanc", "", "Communication"),
            ("Communication Protocols", "DC_COMM_PROTO", r"(CAN\s*2\.0|CAN|RS485|Modbus\s*TCP|Modbus\s*RTU|IEC\s*61850|Ethernet|TCP/IP)", "", "Communication"),
        ]
    elif category == "PCS":
        patterns = [
            # ── Power Ratings ──
            ("Rated AC Power", "PCS_AC_POWER", r"(?:rated|nominal)\s*(?:ac\s*)?power.*?(\d+(?:\.\d+)?)\s*(?:kW|MW)", "kW", "Electrical"),
            ("Max AC Power", "PCS_MAX_AC_POWER", r"(?:max|maximum|peak)\s*(?:ac\s*)?power.*?(\d+(?:\.\d+)?)\s*(?:kW|MW)", "kW", "Electrical"),
            ("Rated DC Power", "PCS_DC_POWER", r"(?:rated|nominal)\s*dc\s*power.*?(\d+(?:\.\d+)?)\s*(?:kW|MW)", "kW", "Electrical"),
            ("Apparent Power", "PCS_APPARENT_POWER", r"(?:apparent|rated)\s*(?:power|capacity).*?(\d+(?:\.\d+)?)\s*(?:kVA|MVA)", "kVA", "Electrical"),
            ("Number of PCS Units", "PCS_UNIT_COUNT", r"(\d+)\s*(?:PCS\s*)?units", "", "Electrical"),
            # ── AC Side ──
            ("AC Voltage", "PCS_AC_VOLTAGE", r"(?:ac|output|grid)\s*(?:side\s*)?voltage.*?(\d+(?:\.\d+)?)\s*(?:kV|V)", "V", "Electrical"),
            ("AC Frequency Range Min", "PCS_FREQ_MIN", r"(?:freq|frequency).*?(\d+(?:\.\d+)?)\s*(?:Hz|hz)\s*(?:to|~|–|-)", "Hz", "Electrical"),
            ("AC Frequency Range Max", "PCS_FREQ_MAX", r"(?:to|~|–|-)\s*(\d+(?:\.\d+)?)\s*(?:Hz|hz)", "Hz", "Electrical"),
            ("AC Output Current", "PCS_AC_CURRENT", r"(?:ac|output|rated)\s*current.*?(\d+(?:\.\d+)?)\s*A", "A", "Electrical"),
            ("Max AC Current", "PCS_MAX_AC_CURRENT", r"(?:max|maximum)\s*(?:ac\s*)?(?:output\s*)?current.*?(\d+(?:\.\d+)?)\s*A", "A", "Electrical"),
            ("Number of Phases", "PCS_PHASES", r"(3[\s-]*phase|three[\s-]*phase|single[\s-]*phase)", "", "Electrical"),
            # ── DC Side ──
            ("DC Voltage Nominal", "PCS_DC_VOLTAGE_NOM", r"(?:dc|battery)\s*(?:nominal\s*)?voltage.*?(\d+(?:\.\d+)?)\s*V", "V", "Electrical"),
            ("DC Voltage Min", "PCS_DC_VOLTAGE_MIN", r"(?:dc|battery|input)\s*(?:voltage\s*)?(?:min|minimum|range).*?(\d+(?:\.\d+)?)\s*V", "V", "Electrical"),
            ("DC Voltage Max", "PCS_DC_VOLTAGE_MAX", r"(?:dc|battery|input)\s*(?:voltage\s*)?(?:max|maximum).*?(\d+(?:\.\d+)?)\s*V", "V", "Electrical"),
            ("Max DC Current", "PCS_MAX_DC_CURRENT", r"(?:max|maximum)\s*(?:dc|input)\s*current.*?(\d+(?:\.\d+)?)\s*A", "A", "Electrical"),
            # ── Performance ──
            ("Peak Efficiency", "PCS_EFFICIENCY", r"(?:max|peak|maximum)\s*(?:conversion\s*)?efficiency.*?(\d+(?:\.\d+)?)\s*%", "%", "Performance"),
            ("Weighted / CEC Efficiency", "PCS_WEIGHTED_EFF", r"(?:weighted|CEC|euro|european)\s*efficiency.*?(\d+(?:\.\d+)?)\s*%", "%", "Performance"),
            ("THD", "PCS_THD", r"THD.*?[<≤]?\s*(\d+(?:\.\d+)?)\s*%", "%", "Performance"),
            ("Individual Harmonic", "PCS_THD_INDIVIDUAL", r"individual\s*(?:harmonic)?.*?[<≤]?\s*(\d+(?:\.\d+)?)\s*%", "%", "Performance"),
            ("Power Factor Range", "PCS_POWER_FACTOR", r"power\s*factor.*?(\d+(?:\.\d+)?)\s*(?:lead|lag|leading|lagging)", "", "Electrical"),
            ("Reactive Power Capability", "PCS_REACTIVE_KVAR", r"reactive\s*(?:power)?\s*(?:capability|capacity|rating).*?(\d+(?:\.\d+)?)\s*(?:kVAr|MVAr)", "kVAr", "Electrical"),
            ("Standby Power Consumption", "PCS_STANDBY_KW", r"standby.*?(?:power|consumption).*?(\d+(?:\.\d+)?)\s*(?:kW|W)", "kW", "Performance"),
            ("Aux Power at Full Load", "PCS_AUX_FULL_KW", r"(?:auxiliary|aux).*?(?:full\s*load|rated).*?(\d+(?:\.\d+)?)\s*(?:kW|W)", "kW", "Performance"),
            # ── Response / Dynamic ──
            ("Response Time (0 to Full)", "PCS_RESPONSE_MS", r"(?:response\s*time|0\s*to\s*full).*?[<≤]?\s*(\d+)\s*(?:ms|millisec)", "ms", "Performance"),
            ("Ramp Rate", "PCS_RAMP_RATE", r"ramp\s*rate.*?(\d+(?:\.\d+)?)\s*(?:MW/s|kW/s|%/s)", "MW/s", "Performance"),
            ("Mode Transition Time", "PCS_MODE_TRANSITION", r"(?:mode\s*transition|charge.*?discharge).*?[<≤]?\s*(\d+)\s*(?:ms|s)", "ms", "Performance"),
            ("Overload Capability", "PCS_OVERLOAD", r"overload.*?(\d+)\s*%.*?(\d+)\s*(?:min|sec|s)", "%", "Performance"),
            # ── Grid Support ──
            ("Black Start Capability", "PCS_BLACK_START", r"black\s*start", "", "Grid Support"),
            ("Primary Frequency Response", "PCS_PFR", r"(?:primary\s*frequency|PFR)", "", "Grid Support"),
            ("Secondary Frequency Response", "PCS_SFR", r"(?:secondary\s*frequency|SFR)", "", "Grid Support"),
            ("Tertiary Frequency Response", "PCS_TFR", r"(?:tertiary\s*frequency|TFR)", "", "Grid Support"),
            ("Reactive Power Compensation", "PCS_REACTIVE", r"reactive\s*power\s*(?:compens|support|control)", "", "Grid Support"),
            ("Voltage Regulation / AVR", "PCS_AVR", r"(?:voltage\s*regulat|AVR\s*mode|automatic\s*voltage)", "", "Grid Support"),
            ("Anti-Islanding", "PCS_ANTI_ISLAND", r"anti[\s-]*island", "", "Grid Support"),
            ("LVRT (Low Voltage Ride Through)", "PCS_LVRT", r"(?:LVRT|low\s*voltage\s*ride[\s-]*through)", "", "Grid Support"),
            ("HVRT (High Voltage Ride Through)", "PCS_HVRT", r"(?:HVRT|high\s*voltage\s*ride[\s-]*through)", "", "Grid Support"),
            ("Frequency Droop Control", "PCS_DROOP", r"(?:frequency\s*)?droop\s*(?:control|setting)", "", "Grid Support"),
            ("Active Power Curtailment", "PCS_CURTAILMENT", r"(?:active\s*power\s*)?curtailment", "", "Grid Support"),
            ("Grid Forming Mode", "PCS_GRID_FORMING", r"grid[\s-]*forming", "", "Grid Support"),
            ("Grid Following Mode", "PCS_GRID_FOLLOWING", r"grid[\s-]*following", "", "Grid Support"),
            ("Grid Synchronization", "PCS_GRID_SYNC", r"(?:grid\s*)?synchroniz.*?(PLL|phase[\s-]*locked)", "", "Grid Support"),
            ("Harmonic Filtering", "PCS_HARMONIC_FILTER", r"(?:harmonic\s*)?filter.*?(active|passive)", "", "Grid Support"),
            # ── Protection ──
            ("AC Side Protections", "PCS_AC_PROTECT", r"(?:ac\s*(?:side\s*)?protection|OV|UV|OF|UF|over[\s-]*voltage|under[\s-]*voltage|over[\s-]*current)", "", "Safety"),
            ("DC Side Protections", "PCS_DC_PROTECT", r"(?:dc\s*(?:side\s*)?protection|reverse\s*polarity|arc\s*fault)", "", "Safety"),
            ("Surge Protection", "PCS_SPD", r"(?:surge\s*protect|SPD).*?(Type\s*[I1]+|Class\s*[I1]+)", "", "Safety"),
            # ── Physical ──
            ("Topology", "PCS_TOPOLOGY", r"(H[\s-]*bridge|NPC|T[\s-]*type|3[\s-]*level|multi[\s-]*level|2[\s-]*level)", "", "Electrical"),
            ("Enclosure Type", "PCS_ENCLOSURE", r"(?:enclosure|housing).*?(indoor|outdoor|container|IP\d+)", "", "Physical"),
            ("Dimensions", "PCS_DIMENSIONS", r"(?:dimension|size).*?(\d+)\s*(?:×|x|X)\s*(\d+)\s*(?:×|x|X)\s*(\d+)\s*mm", "mm", "Physical"),
            ("Weight", "PCS_WEIGHT", r"weight.*?(\d+(?:\.\d+)?)\s*(?:kg|ton|t\b)", "kg", "Physical"),
            ("Protection Rating", "PCS_IP_RATING", r"(IP\d+\w?)", "", "Safety"),
            ("Noise Level", "PCS_NOISE", r"(?:noise|sound).*?[<≤]?\s*(\d+(?:\.\d+)?)\s*dB", "dBA", "Physical"),
            # ── Thermal ──
            ("Cooling Type", "PCS_COOLING", r"(liquid|air|forced\s*air|fan)\s*cool", "", "Thermal"),
            ("Operating Temp Min", "PCS_OP_TEMP_MIN", r"operating\s*temp.*?(-?\d+)\s*[°℃C]", "°C", "Thermal"),
            ("Operating Temp Max", "PCS_OP_TEMP_MAX", r"operating\s*temp.*?(\d+)\s*[°℃C]", "°C", "Thermal"),
            ("Altitude Rating", "PCS_ALTITUDE", r"(?:altitude|elevation).*?[<≤]?\s*(\d[\d,]*)\s*m", "m", "Thermal"),
            ("Seismic Rating", "PCS_SEISMIC", r"(?:seismic|earthquake).*?(?:zone|IS\s*1893).*?(\w+)", "", "Safety"),
            ("Humidity Range", "PCS_HUMIDITY", r"humidity.*?(\d+)\s*%?\s*(?:to|~|–|-)\s*(\d+)\s*%", "%RH", "Thermal"),
            # ── Communication ──
            ("Communication Protocols", "PCS_COMM", r"(Modbus\s*TCP|Modbus\s*RTU|CAN|Ethernet|RS485|TCP/IP|IEC\s*61850|DNP3)", "", "Communication"),
            ("HMI / Local Display", "PCS_HMI", r"(HMI|touch[\s-]*screen|local\s*display)", "", "Communication"),
            ("Remote Monitoring", "PCS_REMOTE", r"remote\s*(?:monitor|access|control)", "", "Communication"),
            ("Data Logging", "PCS_DATA_LOG", r"data\s*log.*?(\d+)\s*(?:year|month|day|GB)", "", "Communication"),
            ("Cybersecurity", "PCS_CYBERSEC", r"(?:cyber[\s-]*security|IEC\s*62443|RBAC)", "", "Communication"),
            # ── Transformer ──
            ("Transformer Type", "PCS_TRAFO_TYPE", r"transformer.*?(dry|oil|cast\s*resin)", "", "Transformer"),
            ("Transformer Rating", "PCS_TRAFO_RATING", r"transformer.*?(?:rating|capacity).*?(\d+(?:\.\d+)?)\s*(?:kVA|MVA)", "kVA", "Transformer"),
            ("Primary Voltage", "PCS_TRAFO_PRIMARY_V", r"(?:primary|LV)\s*(?:side\s*)?voltage.*?(\d+(?:\.\d+)?)\s*V", "V", "Transformer"),
            ("Secondary Voltage", "PCS_TRAFO_SECONDARY_KV", r"(?:secondary|HV|MV)\s*(?:side\s*)?voltage.*?(\d+(?:\.\d+)?)\s*kV", "kV", "Transformer"),
            ("Transformer Impedance", "PCS_TRAFO_IMPEDANCE", r"(?:transformer\s*)?impedance.*?(\d+(?:\.\d+)?)\s*%", "%", "Transformer"),
            ("Transformer Cooling", "PCS_TRAFO_COOLING", r"(?:transformer\s*)?cool.*?(ONAN|ONAF|AN|AF)", "", "Transformer"),
            # ── Standards ──
            ("CEA Grid Code Compliance", "PCS_CEA", r"CEA.*?(?:Grid\s*Code|2019|technical\s*standard)", "", "Standards"),
            ("IEEE 1547 Compliance", "PCS_IEEE1547", r"IEEE\s*1547", "", "Standards"),
            ("IEC 62477 Compliance", "PCS_IEC62477", r"IEC\s*62477", "", "Standards"),
            ("IEC 62109 Compliance", "PCS_IEC62109", r"IEC\s*62109", "", "Standards"),
            ("IEC 61000 EMC", "PCS_IEC61000", r"IEC\s*61000", "", "Standards"),
            ("UL 1741 Compliance", "PCS_UL1741", r"UL\s*1741", "", "Standards"),
            ("Make in India", "PCS_MAKE_IN_INDIA", r"(?:make\s*in\s*india|local\s*content|indigenous).*?(\d+)?\s*%?", "%", "Standards"),
        ]
    elif category == "EMS":
        patterns = [
            # ── Company / Architecture ──
            ("EMS Software Platform", "EMS_PLATFORM", r"(?:EMS|SCADA)\s*(?:software|platform).*?(\w[\w\s]+)", "", "Architecture"),
            ("EMS Architecture", "EMS_ARCHITECTURE", r"(?:EMS\s*)?(?:architecture|topology).*?(centralized|distributed|cloud|edge|hybrid)", "", "Architecture"),
            ("Server Configuration", "EMS_SERVER", r"(redundant|dual|hot[\s-]*standby|single|primary.*?standby)\s*(?:server)?", "", "Architecture"),
            ("Server CPU", "EMS_SERVER_CPU", r"(?:CPU|processor).*?(Intel|AMD|Xeon|Core\s*i\d)", "", "Architecture"),
            ("Server RAM", "EMS_SERVER_RAM", r"(?:RAM|memory).*?(\d+)\s*GB", "GB", "Architecture"),
            ("Server Storage", "EMS_SERVER_STORAGE", r"(?:storage|disk|SSD|HDD).*?(\d+)\s*(?:TB|GB)", "TB", "Architecture"),
            ("Operating System", "EMS_OS", r"(?:operating\s*system|OS).*?(Windows|Linux|RHEL|Ubuntu|CentOS)", "", "Architecture"),
            ("UPS for EMS", "EMS_UPS", r"(?:UPS|backup\s*power).*?(\d+(?:\.\d+)?)\s*(?:kVA|kW)", "kVA", "Architecture"),
            ("UPS Backup Duration", "EMS_UPS_BACKUP", r"(?:UPS|backup).*?(\d+)\s*(?:hour|hr|min)", "hours", "Architecture"),
            # ── Control Functions ──
            ("Active Power Dispatch", "EMS_ACTIVE_DISPATCH", r"(?:active\s*power|MW)\s*(?:dispatch|setpoint|control)", "", "Functionality"),
            ("Reactive Power Control", "EMS_REACTIVE_CTRL", r"(?:reactive\s*power|kVAr)\s*(?:dispatch|setpoint|control)", "", "Functionality"),
            ("AGC Interface", "EMS_AGC", r"(?:AGC|automatic\s*generation\s*control)", "", "Functionality"),
            ("Frequency Regulation (Primary)", "EMS_FREQ_PRIMARY", r"(?:primary\s*frequency|PFR)\s*(?:response|regulation|control)", "", "Functionality"),
            ("Frequency Regulation (Secondary)", "EMS_FREQ_SECONDARY", r"(?:secondary\s*frequency|SFR)\s*(?:response|regulation)", "", "Functionality"),
            ("Frequency Regulation (Tertiary)", "EMS_FREQ_TERTIARY", r"(?:tertiary\s*frequency|TFR)\s*(?:response|regulation)", "", "Functionality"),
            ("Peak Shaving", "EMS_PEAK_SHAVING", r"peak\s*(?:shaving|shave|clipping)", "", "Functionality"),
            ("Load Leveling", "EMS_LOAD_LEVEL", r"load\s*(?:level|leveling|levelling|shifting)", "", "Functionality"),
            ("Ramp Rate Control", "EMS_RAMP_RATE", r"ramp\s*rate\s*(?:control|limit|management)", "", "Functionality"),
            ("SOC Management", "EMS_SOC_MGMT", r"SOC\s*(?:management|optimi|algorithm|predictive)", "", "Functionality"),
            ("Degradation-Aware Dispatch", "EMS_DEG_DISPATCH", r"degradation[\s-]*aware\s*(?:dispatch|control|algorithm)", "", "Functionality"),
            ("Scheduling & Forecasting", "EMS_FORECASTING", r"(?:scheduling|forecast|day[\s-]*ahead|intra[\s-]*day)", "", "Functionality"),
            ("Multi-Unit Coordination", "EMS_MULTI_UNIT", r"(?:multi[\s-]*unit|multi[\s-]*PCS|coordination|orchestrat)", "", "Functionality"),
            ("Black Start Sequence", "EMS_BLACK_START", r"black\s*start\s*(?:logic|sequence|capability|management|support)", "", "Functionality"),
            ("Islanding Detection", "EMS_ISLANDING", r"island(?:ing)?\s*(?:detection|management|mode)", "", "Functionality"),
            ("Arbitrage Mode", "EMS_ARBITRAGE", r"(?:energy\s*)?arbitrage", "", "Functionality"),
            ("RTE Monitoring", "EMS_RTE_MON", r"RTE\s*(?:monitoring|measurement|tracking|report)", "", "Functionality"),
            # ── SCADA ──
            ("SCADA Platform", "EMS_SCADA_PLATFORM", r"SCADA\s*(?:software|platform|system).*?(\w[\w\s]+)", "", "SCADA"),
            ("I/O Points", "EMS_IO_POINTS", r"(?:I/O|IO)\s*points.*?(\d[\d,]*)", "", "SCADA"),
            ("HMI Workstations", "EMS_HMI", r"(?:HMI|operator\s*workstation|human[\s-]*machine).*?(\d+)\s*(?:nos?|units?|screens?)?", "", "SCADA"),
            ("Alarm Management", "EMS_ALARMS", r"alarm\s*(?:management|system|configur|priority)", "", "SCADA"),
            ("SOE Recording", "EMS_SOE", r"(?:SOE|sequence\s*of\s*events?).*?[≤<]?\s*(\d+)\s*ms", "ms", "SCADA"),
            ("Historical Data Storage", "EMS_HISTORIAN", r"(?:historian|historical\s*data|trending|data\s*storage).*?(\d+)\s*(?:year|month|GB|TB)?", "", "SCADA"),
            ("Report Generation", "EMS_REPORTS", r"(?:report\s*generat|daily.*?report|monthly.*?report|annual.*?report)", "", "SCADA"),
            ("Web-Based Access", "EMS_WEB", r"(?:web[\s-]*based|browser[\s-]*based|remote\s*access)", "", "SCADA"),
            ("Mobile App", "EMS_MOBILE", r"(?:mobile\s*app|mobile\s*dashboard|iOS|Android)", "", "SCADA"),
            ("Video Wall", "EMS_VIDEO_WALL", r"(?:video\s*wall|large\s*display|projection)", "", "SCADA"),
            ("Engineering Workstation", "EMS_ENG_WS", r"engineering\s*(?:workstation|station|console)", "", "SCADA"),
            # ── Performance ──
            ("System Availability", "EMS_AVAILABILITY", r"(?:system\s*)?availability.*?[≥>]?\s*(\d+(?:\.\d+)?)\s*%", "%", "Performance"),
            ("Data Refresh Rate", "EMS_REFRESH", r"(?:refresh|scan|update|polling).*?[≤<]?\s*(\d+)\s*(?:s|sec|second|ms)", "s", "Performance"),
            ("MTBF", "EMS_MTBF", r"MTBF.*?(\d[\d,]*)\s*(?:hour|hr)", "hours", "Performance"),
            # ── Communication ──
            ("Protocol to BMS", "EMS_PROTO_BMS", r"(?:to\s*)?BMS.*?(Modbus|CAN|RS485|TCP)", "", "Communication"),
            ("Protocol to PCS", "EMS_PROTO_PCS", r"(?:to\s*)?PCS.*?(Modbus|IEC\s*61850|CAN|TCP)", "", "Communication"),
            ("Protocol to Plant SCADA", "EMS_PROTO_PLANT", r"(?:plant\s*SCADA|DCS).*?(IEC\s*61850|DNP3|Modbus|IEC\s*60870)", "", "Communication"),
            ("Protocol to SLDC/NLDC", "EMS_PROTO_SLDC", r"(?:SLDC|NLDC|RLDC).*?(IEC\s*60870|IEC\s*61850|DNP3)", "", "Communication"),
            ("Metering Interface", "EMS_METERING", r"(?:ABT|special\s*energy\s*)?meter.*?(?:interface|integrat)", "", "Communication"),
            ("GPS Time Sync", "EMS_TIME_SYNC", r"(GPS|NTP|PTP|IEEE\s*1588)\s*(?:based|time|sync)", "", "Communication"),
            ("Redundant Communication", "EMS_REDUNDANT_COMM", r"(?:redundant|dual)\s*(?:communication|Ethernet|fiber|link)", "", "Communication"),
            ("Weather Station", "EMS_WEATHER", r"weather\s*station\s*(?:interface|integrat)", "", "Communication"),
            ("Fiber Optic Network", "EMS_FIBER", r"(?:fiber\s*optic|optical\s*fiber|OFC)\s*(?:network|backbone)", "", "Communication"),
            ("Network Switches", "EMS_SWITCHES", r"(?:network\s*switch|managed\s*switch|industrial\s*switch)", "", "Communication"),
            # ── Cybersecurity ──
            ("IEC 62443 Compliance", "EMS_IEC62443", r"IEC\s*62443.*?(?:SL[\s-]*(\d)|level\s*(\d))?", "", "Security"),
            ("RBAC", "EMS_RBAC", r"(?:RBAC|role[\s-]*based\s*access)", "", "Security"),
            ("Audit Trail", "EMS_AUDIT", r"audit\s*(?:trail|log|record).*?(?:tamper[\s-]*proof)?", "", "Security"),
            ("Encrypted Communication", "EMS_ENCRYPTION", r"(?:encrypted|TLS|SSL|AES)\s*(?:communication|connect|channel)?", "", "Security"),
            ("Firewall", "EMS_FIREWALL", r"(?:firewall|network\s*segmentat)", "", "Security"),
            ("Intrusion Detection", "EMS_IDS", r"(?:IDS|intrusion\s*detect|intrusion\s*prevent)", "", "Security"),
            ("Patch Management", "EMS_PATCH", r"(?:patch\s*management|vulnerability\s*assess)", "", "Security"),
            # ── Software Origin ──
            ("Make in India", "EMS_ORIGIN", r"(indigenous|Make\s*in\s*India|developed\s*in\s*India|Indian\s*(?:software|OEM))", "", "Compliance"),
            # ── Testing ──
            ("FAT Plan", "EMS_FAT", r"(?:FAT|factory\s*acceptance\s*test)", "", "Testing"),
            ("SAT Plan", "EMS_SAT", r"(?:SAT|site\s*acceptance\s*test)", "", "Testing"),
            ("IEC 61850 Conformance", "EMS_IEC61850_CONF", r"IEC\s*61850\s*conformance", "", "Testing"),
        ]
    else:
        patterns = []

    for name, code, pattern, unit, section in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            try:
                value = match.group(1).replace(",", "")
            except (IndexError, AttributeError):
                value = match.group(0).strip()
            extracted.append({
                "name": name, "code": code, "value": value,
                "unit": unit, "section": section, "status": "pass", "confidence": 0.85,
            })

    # Check certifications (all categories)
    certs = []
    cert_checks = [
        ("iec 62619", "IEC 62619"), ("ul 1973", "UL 1973"), ("ul 9540a", "UL 9540A"),
        ("ul 9540", "UL 9540"), ("un 38.3", "UN 38.3"), ("un38.3", "UN 38.3"),
        ("iec 62933", "IEC 62933"), ("iec 62477", "IEC 62477"), ("iec 61000", "IEC 61000"),
        ("iec 61850", "IEC 61850"), ("ieee 1547", "IEEE 1547"), ("nfpa", "NFPA"),
        ("bis", "BIS"), ("cea", "CEA"), ("cerc", "CERC"),
    ]
    for keyword, cert_name in cert_checks:
        if keyword in text_lower:
            certs.append(cert_name)
    if certs:
        extracted.append({"name": "Certifications", "code": f"{category.upper().replace(' ', '_')}_CERTS",
                         "value": ", ".join(certs), "unit": "", "section": "Safety", "status": "pass", "confidence": 0.95})

    # Chemistry / Technology (for all categories)
    chemistry_map = {"lfp": "LFP", "lifepo4": "LFP", "nmc": "NMC", "lto": "LTO", "sodium": "Sodium-ion"}
    for kw, val in chemistry_map.items():
        if kw in text_lower:
            extracted.append({"name": "Chemistry", "code": f"{category.upper().replace(' ', '_')}_CHEMISTRY",
                             "value": val, "unit": "", "section": "General", "status": "pass", "confidence": 0.92})
            break

    # Form Factor
    form_factors = {"prismatic": "Prismatic", "cylindrical": "Cylindrical", "pouch": "Pouch", "blade": "Blade"}
    for kw, val in form_factors.items():
        if kw in text_lower:
            extracted.append({"name": "Form Factor", "code": f"{category.upper().replace(' ', '_')}_FORM",
                             "value": val, "unit": "", "section": "Physical", "status": "pass", "confidence": 0.90})
            break

    # Model Number — try to find it
    model_match = re.search(r"(?:model|type|part)\s*(?:no|number|#)?[\s:]*([A-Z0-9][\w\-]{3,})", text, re.IGNORECASE)
    if model_match:
        extracted.append({"name": "Model Number", "code": f"{category.upper().replace(' ', '_')}_MODEL",
                         "value": model_match.group(1), "unit": "", "section": "General", "status": "pass", "confidence": 0.88})

    # Manufacturer
    mfr_match = re.search(r"(?:manufacturer|brand|made\s*by|company)[\s:]*([A-Z][\w\s]{2,30})", text, re.IGNORECASE)
    if mfr_match:
        extracted.append({"name": "Manufacturer", "code": f"{category.upper().replace(' ', '_')}_MFR",
                         "value": mfr_match.group(1).strip(), "unit": "", "section": "General", "status": "pass", "confidence": 0.85})

    # Design Life
    life_match = re.search(r"(?:design|service|expected)\s*life.*?(\d+)\s*years?", text, re.IGNORECASE)
    if life_match:
        extracted.append({"name": "Design Life", "code": f"{category.upper().replace(' ', '_')}_DESIGN_LIFE",
                         "value": life_match.group(1), "unit": "years", "section": "Performance", "status": "pass", "confidence": 0.90})

    # Warranty
    warranty_match = re.search(r"warranty.*?(\d+)\s*years?", text, re.IGNORECASE)
    if warranty_match:
        extracted.append({"name": "Warranty Period", "code": f"{category.upper().replace(' ', '_')}_WARRANTY",
                         "value": warranty_match.group(1), "unit": "years", "section": "General", "status": "pass", "confidence": 0.88})

    return extracted


def _check_compliance(params: list, category: str) -> list:
    """
    Compare extracted values against BESS industry-standard thresholds.
    Marks each parameter as pass / fail / info based on real requirements.
    This is what makes compliance scoring meaningful — not everything is "pass".
    """

    # Thresholds: code → (operator, threshold, description)
    # "gte" = value must be >= threshold to pass
    # "lte" = value must be <= threshold to pass
    # "eq"  = value must match to pass
    # "has" = value must contain keyword to pass
    THRESHOLDS = {
        # ── Cell ──
        "CELL_CAPACITY_AH":       ("gte", 100,   "≥100 Ah for utility-scale"),
        "CELL_VOLTAGE_V":         ("gte", 3.0,   "≥3.0V nominal for LFP/NMC"),
        "CELL_CYCLE_LIFE":        ("gte", 6000,  "≥6000 cycles per CERC/NTPC norms"),
        "CELL_CALENDAR_LIFE":     ("gte", 15,    "≥15 years calendar life"),
        "CELL_RTE":               ("gte", 90,    "≥90% round-trip efficiency"),
        "CELL_COULOMBIC_EFF":     ("gte", 99,    "≥99% coulombic efficiency"),
        "CELL_EOL_RETENTION":     ("gte", 70,    "≥70% EOL capacity retention"),
        "CELL_SELF_DISCHARGE":    ("lte", 3,     "≤3% per month self-discharge"),
        "CELL_CHARGE_TEMP_MIN":   ("lte", 0,     "≤0°C charge temp min"),
        "CELL_DISCHARGE_TEMP_MIN":("lte", -10,   "≤-10°C discharge temp min"),
        "CELL_CHARGE_TEMP_MAX":   ("gte", 45,    "≥45°C charge temp max"),
        "CELL_DISCHARGE_TEMP_MAX":("gte", 50,    "≥50°C discharge temp max"),
        "CELL_ENERGY_DENSITY_WH_KG": ("gte", 140, "≥140 Wh/kg for modern LFP"),
        "CELL_DEGRADATION":       ("lte", 3,     "≤3% annual degradation"),
        "CELL_TR_ONSET":          ("gte", 200,   "≥200°C thermal runaway onset"),
        # ── DC Block / Container ──
        "DC_RATED_ENERGY":        ("gte", 1000,  "≥1000 kWh per container"),
        "DC_RTE":                 ("gte", 90,    "≥90% DC round-trip efficiency"),
        "DC_SOC_ACC":             ("lte", 5,     "≤5% SOC accuracy"),
        "DC_COOLING_KW":          ("gte", 30,    "≥30 kW cooling capacity"),
        "DC_OP_TEMP_MIN":         ("lte", -10,   "≤-10°C operating temp min"),
        "DC_OP_TEMP_MAX":         ("gte", 45,    "≥45°C operating temp max"),
        "DC_ALTITUDE":            ("gte", 1000,  "≥1000m altitude rating"),
        "DC_CYCLE_LIFE":          ("gte", 6000,  "≥6000 cycles"),
        "DC_CALENDAR_LIFE":       ("gte", 15,    "≥15 years"),
        # ── PCS ──
        "PCS_EFFICIENCY":         ("gte", 98,    "≥98% peak efficiency"),
        "PCS_WEIGHTED_EFF":       ("gte", 97,    "≥97% weighted/CEC efficiency"),
        "PCS_THD":                ("lte", 5,     "≤5% THD at rated output"),
        "PCS_THD_INDIVIDUAL":     ("lte", 3,     "≤3% individual harmonic"),
        "PCS_RESPONSE_MS":        ("lte", 200,   "≤200ms 0-to-full response"),
        "PCS_NOISE":              ("lte", 80,    "≤80 dBA noise level"),
        "PCS_OP_TEMP_MIN":        ("lte", 0,     "≤0°C operating temp min"),
        "PCS_OP_TEMP_MAX":        ("gte", 45,    "≥45°C operating temp max"),
        "PCS_STANDBY_KW":         ("lte", 2,     "≤2 kW standby consumption"),
        # ── EMS ──
        "EMS_AVAILABILITY":       ("gte", 99,    "≥99% system availability"),
        "EMS_REFRESH":            ("lte", 2,     "≤2 second data refresh"),
        "EMS_SOE":                ("lte", 1,     "≤1ms SOE resolution"),
    }

    # Certifications that are REQUIRED — varies by category
    REQUIRED_CERTS_MAP = {
        "Cell":     {"IEC 62619", "UL 1973", "UL 9540A", "UN 38.3"},
        "DC Block": {"UL 9540", "UL 9540A"},
        "PCS":      {"IEC 62477", "IEC 61000"},
        "EMS":      {"IEC 61850"},
    }
    REQUIRED_CERTS = REQUIRED_CERTS_MAP.get(category, set())

    for p in params:
        code = p.get("code", "")
        value_str = p.get("value", "")

        # Check numeric thresholds
        if code in THRESHOLDS:
            operator, threshold, desc = THRESHOLDS[code]
            try:
                val = float(value_str.replace(",", "").replace("≥", "").replace("≤", "").replace(">", "").replace("<", "").strip())
                if operator == "gte":
                    p["status"] = "pass" if val >= threshold else "fail"
                elif operator == "lte":
                    p["status"] = "pass" if val <= threshold else "fail"
                p["threshold"] = f"{'≥' if operator == 'gte' else '≤'}{threshold}"
                p["threshold_desc"] = desc
            except (ValueError, TypeError):
                p["status"] = "info"  # Can't compare, just informational

        # Check certifications
        elif code.endswith("_CERTS"):
            certs_found = set(c.strip() for c in value_str.split(","))
            missing = REQUIRED_CERTS - certs_found
            if not missing:
                p["status"] = "pass"
                p["threshold_desc"] = "All required certs present"
            else:
                p["status"] = "fail"
                p["threshold_desc"] = f"Missing: {', '.join(missing)}"

        # Chemistry check
        elif code.endswith("_CHEMISTRY"):
            p["status"] = "pass" if value_str.upper() in ("LFP", "LFEPO4") else "info"
            p["threshold_desc"] = "LFP preferred per CEA/CERC"

        # Form factor
        elif code.endswith("_FORM"):
            p["status"] = "pass" if value_str.lower() == "prismatic" else "info"
            p["threshold_desc"] = "Prismatic preferred for utility BESS"

        # IP rating check
        elif code.endswith("_IP_RATING"):
            try:
                ip_num = int(re.search(r"(\d{2})", value_str).group(1))
                p["status"] = "pass" if ip_num >= 54 else "fail"
                p["threshold"] = "≥IP54"
                p["threshold_desc"] = "≥IP54 for outdoor installations"
            except (AttributeError, ValueError):
                p["status"] = "info"

        # Grid support features — presence = pass
        elif code in ("PCS_BLACK_START", "PCS_PFR", "PCS_SFR", "PCS_TFR",
                      "PCS_REACTIVE", "PCS_AVR", "PCS_ANTI_ISLAND",
                      "PCS_LVRT", "PCS_HVRT", "PCS_GRID_FORMING",
                      "PCS_GRID_FOLLOWING", "PCS_DROOP", "PCS_CURTAILMENT"):
            p["status"] = "pass"
            p["threshold_desc"] = "Feature present"

        # EMS features — presence = pass
        elif code.startswith("EMS_") and not code.endswith(("_AVAILABILITY", "_REFRESH", "_SOE", "_MTBF")):
            p["status"] = "pass"
            p["threshold_desc"] = "Feature present"

        # Everything else stays as-is or gets "info" (no threshold to check)
        else:
            if p.get("status") == "pass":
                p["status"] = "info"  # No threshold → just informational, not a real "pass"

    return params


# Required certifications to check for missing ones
_REQUIRED_CERTS_BY_CATEGORY = {
    "Cell": [
        ("IEC 62619", "IEC 62619 — Cell Safety"),
        ("UL 1973", "UL 1973 — Stationary Batteries"),
        ("UL 9540A", "UL 9540A — Thermal Runaway Fire Test"),
        ("UN 38.3", "UN 38.3 — Transport Safety"),
    ],
    "DC Block": [
        ("UL 9540", "UL 9540 — ESS System Level"),
        ("UL 9540A", "UL 9540A — Container-Level Fire Test"),
        ("IEC 62619", "IEC 62619 — Cell Safety"),
        ("NFPA", "NFPA 855 — ESS Fire Safety"),
    ],
    "PCS": [
        ("IEC 62477", "IEC 62477 — Power Electronics"),
        ("IEC 61000", "IEC 61000 — EMC Compliance"),
        ("IEEE 1547", "IEEE 1547 — Grid Interconnection"),
    ],
    "EMS": [
        ("IEC 61850", "IEC 61850 — Substation Communication"),
        ("IEC 62443", "IEC 62443 — Cybersecurity"),
    ],
}


def _add_missing_required(params: list, category: str) -> list:
    """Check for required certifications and specs that are MISSING from the datasheet.
    In real life, missing required certs = fail."""

    existing_codes = {p["code"] for p in params}
    existing_values = " ".join(p.get("value", "") for p in params).lower()

    # Check required certs
    for cert_keyword, cert_name in _REQUIRED_CERTS_BY_CATEGORY.get(category, []):
        if cert_keyword.lower() not in existing_values:
            params.append({
                "name": cert_name,
                "code": f"MISSING_{cert_keyword.replace(' ', '_').replace('.', '').upper()}",
                "value": "NOT FOUND in datasheet",
                "unit": "",
                "section": "Missing / Required",
                "status": "fail",
                "confidence": 1.0,
                "threshold_desc": f"Required for Indian BESS — not found in document",
            })

    # Check for critical missing specs by category
    critical_specs = {
        "Cell": [
            ("CELL_CYCLE_LIFE", "Cycle Life", "cycles", "≥6000 cycles required"),
            ("CELL_RTE", "Round-Trip Efficiency", "%", "≥90% required"),
            ("CELL_CAPACITY_AH", "Nominal Capacity", "Ah", "Must be specified"),
            ("CELL_VOLTAGE_V", "Nominal Voltage", "V", "Must be specified"),
        ],
        "DC Block": [
            ("DC_RATED_ENERGY", "Rated Energy", "kWh", "Must be specified"),
            ("DC_COOLING", "Cooling Type", "", "Must be specified"),
        ],
        "PCS": [
            ("PCS_EFFICIENCY", "Peak Efficiency", "%", "≥98% required"),
            ("PCS_AC_POWER", "Rated AC Power", "kW", "Must be specified"),
            ("PCS_THD", "THD", "%", "≤5% required"),
        ],
        "EMS": [
            ("EMS_AVAILABILITY", "System Availability", "%", "≥99% required"),
        ],
    }

    for code, name, unit, desc in critical_specs.get(category, []):
        if code not in existing_codes:
            params.append({
                "name": name,
                "code": f"MISSING_{code}",
                "value": "NOT FOUND in datasheet",
                "unit": unit,
                "section": "Missing / Required",
                "status": "fail",
                "confidence": 1.0,
                "threshold_desc": desc,
            })

    return params


def extract_from_datasheet(contents: bytes, filename: str, category: str) -> list:
    """Extract specs from datasheet, check compliance against industry thresholds,
    and flag missing required parameters."""
    text = extract_text_from_file(contents, filename)
    if not text:
        return []

    print(f"[Datasheet] Keyword extraction for {category} ({len(text)} chars)...")
    results = extract_specs_keyword(text, category)

    # Real compliance checking — compare values against thresholds
    results = _check_compliance(results, category)

    # Flag missing required specs/certs
    results = _add_missing_required(results, category)

    pass_count = sum(1 for p in results if p.get("status") == "pass")
    fail_count = sum(1 for p in results if p.get("status") == "fail")
    info_count = sum(1 for p in results if p.get("status") == "info")
    print(f"[Datasheet] {len(results)} specs: {pass_count} pass, {fail_count} fail, {info_count} info")
    return results
