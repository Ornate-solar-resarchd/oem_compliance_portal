"""
RFQ endpoints — Upload any BESS RFQ document (PDF/Excel) and extract technical requirements.
Uses keyword-based extraction from actual document text. No hardcoded templates.
"""
from fastapi import APIRouter, UploadFile, File, Form
from pydantic import BaseModel
from typing import List, Optional
import io
from app.data.seed import RFQS, COMPONENTS, PARAMETERS

router = APIRouter(prefix="/rfq")


# ─── PDF / Excel Text Extraction ───

def _extract_text_from_pdf(contents: bytes) -> str:
    """Extract text from PDF bytes using PyPDF2."""
    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(io.BytesIO(contents))
        text_parts = []
        for page in reader.pages:
            page_text = page.extract_text()
            if page_text:
                text_parts.append(page_text)
        return "\n".join(text_parts)
    except Exception as e:
        print(f"PDF extraction error: {e}")
        return ""


def _extract_text_from_excel(contents: bytes) -> str:
    """Extract text from Excel bytes using openpyxl."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(contents), data_only=True)
        text_parts = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                row_text = " ".join(str(cell) for cell in row if cell is not None)
                if row_text.strip():
                    text_parts.append(row_text)
        return "\n".join(text_parts)
    except Exception as e:
        print(f"Excel extraction error: {e}")
        return ""


def _extract_text(contents: bytes, filename: str) -> str:
    """Extract text from uploaded file based on extension."""
    ext = (filename or "").rsplit(".", 1)[-1].lower()
    if ext == "pdf":
        return _extract_text_from_pdf(contents)
    elif ext in ("xlsx", "xls"):
        return _extract_text_from_excel(contents)
    elif ext == "csv":
        try:
            return contents.decode("utf-8", errors="ignore")
        except:
            return ""
    return ""


def _extract_project_meta(text: str) -> dict:
    """Extract project-level metadata from document text."""
    import re
    meta = {}

    # Capacity patterns
    mw_match = re.search(r"(\d+)\s*MW\s*/\s*(\d+)\s*MWh", text)
    if mw_match:
        meta["capacity"] = f"{mw_match.group(1)} MW / {mw_match.group(2)} MWh"

    # Location
    locations = re.findall(r"(?:Tamil Nadu|Rajasthan|Bihar|Kerala|Maharashtra|Gujarat|Uttar Pradesh|Madhya Pradesh|Karnataka|Andhra Pradesh|Telangana|Haryana|Punjab|West Bengal|Odisha|Jharkhand|Chhattisgarh|Bikaner|Chennai|Barh|Patna|Sitapur|Kochi)", text, re.IGNORECASE)
    if locations:
        meta["location"] = f"{locations[0]}, India"

    # RFQ reference
    ref_match = re.search(r"(?:RFQ|IFB|Tender)\s*(?:Reference|No\.?|Number)[\s:]*([A-Z0-9\-/]+)", text)
    if ref_match:
        meta["rfq_ref"] = ref_match.group(1)

    # Solar PV
    solar_match = re.search(r"(\d+)\s*MWp", text)
    if solar_match:
        meta["solar_pv"] = f"{solar_match.group(1)} MWp"

    # Design life
    life_match = re.search(r"design\s*life.*?(\d+)\s*years", text, re.IGNORECASE)
    if life_match:
        meta["design_life"] = f"{life_match.group(1)} Years"

    # Issuer
    issuers = re.findall(r"(NTPC|Evolve Energy|BSPGCL|TNGECL|Eagle Infra|NHPC|SECI|PGCIL)", text, re.IGNORECASE)
    if issuers:
        meta["issuer"] = issuers[0].upper()

    return meta


# ─── Endpoints ───

@router.get("/")
async def list_rfqs():
    return {"items": RFQS, "total": len(RFQS), "page": 1, "per_page": 20}


@router.post("/upload-multi")
async def upload_rfq_multi(
    files: List[UploadFile] = File(...),
    customer_name: str = Form(...),
    project_name: str = Form(...),
):
    """Upload multiple compliance sheets (Battery, PCS, EMS, HVAC, etc.) into ONE RFQ.
    All files are parsed, extracted, merged, and deduplicated into a single RFQ entry."""

    all_requirements = []
    all_compliance = {}
    all_text = ""
    file_names = []
    total_size = 0

    for f in files:
        contents = await f.read()
        total_size += len(contents)
        file_names.append(f.filename or "unknown")

        # Extract text
        doc_text = _extract_text(contents, f.filename or "")
        all_text += " " + doc_text

        # Extract compliance data
        from app.data.compliance_extraction import extract_compliance_from_rfq
        compliance = extract_compliance_from_rfq(doc_text)
        for cat, cat_data in compliance.items():
            if cat not in all_compliance:
                all_compliance[cat] = cat_data
            else:
                # Merge: update found values
                for i, sheet in enumerate(cat_data.get("sheets", [])):
                    if i < len(all_compliance[cat].get("sheets", [])):
                        existing_codes = {p["code"] for p in all_compliance[cat]["sheets"][i].get("parameters", [])}
                        for p in sheet.get("parameters", []):
                            if p["code"] not in existing_codes:
                                all_compliance[cat]["sheets"][i]["parameters"].append(p)
                            elif p.get("found"):
                                # Update with found value
                                for ep in all_compliance[cat]["sheets"][i]["parameters"]:
                                    if ep["code"] == p["code"] and not ep.get("found"):
                                        ep.update(p)
                                        break

        # Extract general requirements
        from app.data.rfq_extraction import extract_from_text
        general_reqs = extract_from_text(doc_text)
        all_requirements.extend(general_reqs)

    # Flatten ALL compliance parameters into requirements (not just found ones)
    compliance_reqs = []
    for category, cat_data in all_compliance.items():
        for sheet in cat_data.get("sheets", []):
            for p in sheet.get("parameters", []):
                compliance_reqs.append({
                    "parameter": p["parameter"],
                    "code": p["code"],
                    "required_value": p.get("rfq_requirement", "—"),
                    "unit": "",
                    "section": f"{category} — {sheet['name']}",
                    "found": p.get("found", False),
                })

    # Merge and deduplicate
    seen_codes = set()
    final_reqs = []
    for r in compliance_reqs + all_requirements:
        code = r.get("code", "")
        if code and code in seen_codes:
            continue
        if code:
            seen_codes.add(code)
        final_reqs.append(r)

    # Extract project metadata from combined text
    project_meta = _extract_project_meta(all_text)
    if not project_meta.get("issuer"):
        project_meta["issuer"] = customer_name

    # Create single RFQ
    new_id = f"rfq-{len(RFQS) + 1:03d}"
    new_rfq = {
        "id": new_id,
        "customer_name": customer_name,
        "project_name": project_name,
        "status": "active",
        "created_at": "2026-03-30T10:00:00Z",
        "requirements": final_reqs,
        "source_file": ", ".join(file_names),
        "file_names": file_names,
        "files_count": len(files),
        "file_size": total_size,
        "file_type": "multi",
        "extraction_method": "compliance_template_multi",
        "text_extracted": len(all_text) > 0,
        "text_length": len(all_text),
        "compliance_data": all_compliance,
        "gdrive_url": "",
    }
    if project_meta:
        new_rfq["project_meta"] = project_meta
    RFQS.append(new_rfq)

    return {
        "status": "extracted",
        "rfq_id": new_id,
        "customer_name": customer_name,
        "project_name": project_name,
        "file_names": file_names,
        "files_processed": len(files),
        "total_size_bytes": total_size,
        "text_extracted": len(all_text) > 0,
        "text_length": len(all_text),
        "requirements_extracted": len(final_reqs),
        "requirements": final_reqs,
        "project_meta": project_meta,
        "message": f"Extracted {len(final_reqs)} unique parameters from {len(files)} compliance sheet(s)",
    }


@router.post("/upload")
async def upload_rfq(
    file: UploadFile = File(...),
    customer_name: str = Form(...),
    project_name: str = Form(...),
):
    """Upload any RFQ document (PDF/Excel/CSV) and extract technical requirements.

    The system:
    1. Reads the actual document text (PDF → PyPDF2, Excel → openpyxl)
    2. Scans for 50+ BESS technical parameters using keyword matching
    3. Extracts values (capacities, percentages, certifications, etc.)
    4. Returns structured requirements for OEM comparison
    """
    contents = await file.read()
    file_size = len(contents)
    file_ext = (file.filename or "").rsplit(".", 1)[-1].lower()

    gdrive_url = ""

    # Step 1: Extract text from the document
    document_text = _extract_text(contents, file.filename or "")
    text_length = len(document_text)

    # Step 2: Extract using compliance templates (344 params) + RFQ extraction
    from app.data.compliance_extraction import extract_compliance_from_rfq
    from app.data.rfq_extraction import extract_from_text

    # Get compliance-template-based extraction (344 params across 5 categories)
    compliance_data = extract_compliance_from_rfq(document_text)

    # Flatten ALL compliance data into requirements list (show everything)
    extracted_requirements = []
    for category, cat_data in compliance_data.items():
        for sheet in cat_data.get("sheets", []):
            for p in sheet.get("parameters", []):
                extracted_requirements.append({
                    "parameter": p["parameter"],
                    "code": p["code"],
                    "required_value": p.get("rfq_requirement", "—"),
                    "unit": "",
                    "section": f"{category} — {sheet['name']}",
                    "found": p.get("found", False),
                })

    # Also run the general RFQ extraction for anything the templates missed
    general_reqs = extract_from_text(document_text)
    existing_codes = {r["code"] for r in extracted_requirements}
    for req in general_reqs:
        if req.get("code") not in existing_codes:
            extracted_requirements.append(req)

    # Step 3: Extract project metadata
    project_meta = _extract_project_meta(document_text)
    if not project_meta.get("issuer"):
        project_meta["issuer"] = customer_name

    # Step 4: Create and store the RFQ
    new_id = f"rfq-{len(RFQS) + 1:03d}"
    new_rfq = {
        "id": new_id,
        "customer_name": customer_name,
        "project_name": project_name,
        "status": "active",
        "created_at": "2026-03-19T10:00:00Z",
        "requirements": extracted_requirements,
        "source_file": file.filename,
        "file_size": file_size,
        "file_type": file_ext,
        "extraction_method": "compliance_template",
        "text_extracted": text_length > 0,
        "text_length": text_length,
        "compliance_data": compliance_data,
        "gdrive_url": gdrive_url,
    }
    if project_meta:
        new_rfq["project_meta"] = project_meta
    RFQS.append(new_rfq)

    return {
        "status": "extracted",
        "rfq_id": new_id,
        "customer_name": customer_name,
        "project_name": project_name,
        "file_name": file.filename,
        "file_size_bytes": file_size,
        "text_extracted": text_length > 0,
        "text_length": text_length,
        "requirements_extracted": len(extracted_requirements),
        "requirements": extracted_requirements,
        "project_meta": project_meta,
        "gdrive_url": gdrive_url,
        "gdrive_saved": bool(gdrive_url),
        "message": f"Extracted {len(extracted_requirements)} technical requirements from '{file.filename}' ({text_length} chars parsed)",
    }


@router.get("/{rfq_id}")
async def get_rfq(rfq_id: str):
    rfq = next((r for r in RFQS if r["id"] == rfq_id), None)
    if not rfq:
        return {"error": "RFQ not found"}
    return rfq


@router.get("/{rfq_id}/match")
async def match_rfq(rfq_id: str):
    rfq = next((r for r in RFQS if r["id"] == rfq_id), None)
    if not rfq:
        return {"error": "RFQ not found"}

    results = []
    for comp in COMPONENTS:
        params = PARAMETERS.get(comp["id"], [])
        if not params:
            continue

        param_map = {p["code"]: p for p in params}
        total_reqs = len(rfq["requirements"])
        passed = 0
        details = []

        for req in rfq["requirements"]:
            oem_param = param_map.get(req["code"])
            if not oem_param:
                details.append({
                    "parameter": req["parameter"],
                    "required": req["required_value"],
                    "oem_value": "N/A",
                    "match": False,
                    "section": req.get("section", ""),
                })
                continue

            oem_val = oem_param["value"]
            required = req["required_value"]
            match = False

            if required.startswith(">="):
                try:
                    match = float(oem_val) >= float(required[2:])
                except (ValueError, TypeError):
                    match = False
            elif required.startswith("<="):
                try:
                    match = float(oem_val) <= float(required[2:])
                except (ValueError, TypeError):
                    match = False
            elif required.lower() in ("yes", "true", "required"):
                match = oem_val.lower() in ("yes", "true", "certified", "pass")
            else:
                match = (
                    oem_val.strip().lower() == required.strip().lower()
                    or required.lower() in oem_val.lower()
                )

            if match:
                passed += 1

            details.append({
                "parameter": req["parameter"],
                "required": required,
                "oem_value": f"{oem_val} {oem_param['unit']}".strip(),
                "match": match,
                "section": req.get("section", ""),
            })

        match_pct = round((passed / total_reqs) * 100, 1) if total_reqs > 0 else 0
        results.append({
            "component_id": comp["id"],
            "model_name": comp["model_name"],
            "oem_name": comp["oem_name"],
            "match_percentage": match_pct,
            "passed": passed,
            "total": total_reqs,
            "details": details,
        })

    results.sort(key=lambda x: x["match_percentage"], reverse=True)
    return {"rfq_id": rfq_id, "customer": rfq["customer_name"], "matches": results}


@router.get("/{rfq_id}/compliance")
async def get_rfq_compliance(rfq_id: str, categories: str = ""):
    """Generate compliance sheets from an RFQ using the standard 344-parameter template.

    Returns structured data matching the 5 compliance sheet templates:
    - Battery (Cell & Module, Container, BMS, Degradation)
    - PCS Technical Data
    - EMS/SCADA Technical Data
    - HVAC & Fire Safety
    - Functional Guarantees, Spares, References
    """
    rfq = next((r for r in RFQS if r["id"] == rfq_id), None)
    if not rfq:
        return {"error": "RFQ not found"}

    # Get the original document text from stored RFQ
    # Rebuild text from requirements if original text not available
    doc_text = ""
    if rfq.get("requirements"):
        # Build searchable text from extracted requirements
        parts = []
        for req in rfq["requirements"]:
            parts.append(f"{req.get('parameter', '')} {req.get('required_value', '')} {req.get('unit', '')}")
        doc_text = " ".join(parts)

    # Also add project metadata
    meta = rfq.get("project_meta", {})
    for k, v in meta.items():
        doc_text += f" {v}"

    # Parse requested categories
    cat_list = [c.strip() for c in categories.split(",")] if categories else None

    from app.data.compliance_extraction import extract_compliance_from_rfq
    compliance_data = extract_compliance_from_rfq(doc_text, cat_list)

    return {
        "rfq_id": rfq_id,
        "customer": rfq["customer_name"],
        "project": rfq.get("project_name", ""),
        "compliance_sheets": compliance_data,
        "template_version": "1.0",
        "total_categories": len(compliance_data),
    }


@router.get("/compliance-templates")
async def get_compliance_templates():
    """Return all available compliance sheet templates with their parameters."""
    from app.data.compliance_extraction import get_all_template_parameters
    params = get_all_template_parameters()

    # Group by category
    by_category = {}
    for p in params:
        cat = p["category"]
        if cat not in by_category:
            by_category[cat] = {"sheets": {}, "total": 0}
        sheet = p["sheet_name"]
        if sheet not in by_category[cat]["sheets"]:
            by_category[cat]["sheets"][sheet] = []
        by_category[cat]["sheets"][sheet].append(p)
        by_category[cat]["total"] += 1

    return {
        "templates": by_category,
        "total_parameters": len(params),
        "categories": list(by_category.keys()),
    }


@router.post("/upload-compliance")
async def upload_compliance_sheet(
    file: UploadFile = File(...),
    rfq_id: str = Form(""),
    customer_name: str = Form(""),
):
    """Upload a filled compliance sheet (Excel) and extract OEM responses.

    This reads the standard compliance Excel template with OEM-filled data
    and extracts all responses, compliance status, and remarks.
    """
    contents = await file.read()
    file_ext = (file.filename or "").rsplit(".", 1)[-1].lower()

    if file_ext not in ("xlsx", "xls"):
        return {"error": "Only Excel files (.xlsx) are supported for compliance sheets"}

    import openpyxl
    import io

    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    results = {}

    for sn in wb.sheetnames:
        ws = wb[sn]
        sheet_params = []
        current_section = ""

        for row in ws.iter_rows(min_row=5, values_only=False):
            vals = [c.value for c in row[:6]]
            sno, param, requirement, response, compliance, remarks = (vals + [None]*6)[:6]

            # Section headers
            if param and not sno and str(param).strip():
                current_section = str(param).strip()
                continue

            # Parameter rows with data
            if sno and param:
                sheet_params.append({
                    "code": str(sno).strip(),
                    "parameter": str(param).strip(),
                    "rfq_requirement": str(requirement).strip() if requirement else "—",
                    "oem_response": str(response).strip() if response else "",
                    "compliance": str(compliance).strip() if compliance else "",
                    "remarks": str(remarks).strip() if remarks else "",
                    "section": current_section,
                    "has_response": bool(response),
                })

        if sheet_params:
            filled = sum(1 for p in sheet_params if p["has_response"])
            compliant = sum(1 for p in sheet_params if str(p["compliance"]).upper() == "Y")
            results[sn] = {
                "parameters": sheet_params,
                "total": len(sheet_params),
                "filled": filled,
                "fill_rate": round((filled / len(sheet_params)) * 100, 1) if sheet_params else 0,
                "compliant": compliant,
                "compliance_rate": round((compliant / len(sheet_params)) * 100, 1) if sheet_params else 0,
            }

    return {
        "status": "extracted",
        "file_name": file.filename,
        "sheets": results,
        "total_sheets": len(results),
        "total_parameters": sum(s["total"] for s in results.values()),
        "total_filled": sum(s["filled"] for s in results.values()),
        "total_compliant": sum(s["compliant"] for s in results.values()),
    }


class RFQCreate(BaseModel):
    customer_name: str
    project_name: str
    requirements: List[dict]


@router.post("/")
async def create_rfq(rfq: RFQCreate):
    new_id = f"rfq-{len(RFQS) + 1:03d}"
    new_rfq = {
        "id": new_id,
        "customer_name": rfq.customer_name,
        "project_name": rfq.project_name,
        "status": "active",
        "created_at": "2026-03-19T10:00:00Z",
        "requirements": rfq.requirements,
    }
    RFQS.append(new_rfq)
    return new_rfq
