"""
RFQ endpoints — Upload any BESS RFQ document (PDF/Excel) and extract technical requirements.
Uses keyword-based extraction from actual document text. No hardcoded templates.
"""
from fastapi import APIRouter, UploadFile, File, Form
from pydantic import BaseModel
from typing import List, Optional
import io
from app.data.seed import RFQS, COMPONENTS, PARAMETERS

router = APIRouter(prefix="/rfq")


# ─── PDF / Excel Text Extraction ───

def _extract_text_from_pdf(contents: bytes) -> str:
    """Extract text from PDF bytes using PyPDF2."""
    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(io.BytesIO(contents))
        text_parts = []
        for page in reader.pages:
            page_text = page.extract_text()
            if page_text:
                text_parts.append(page_text)
        return "\n".join(text_parts)
    except Exception as e:
        print(f"PDF extraction error: {e}")
        return ""


def _extract_text_from_excel(contents: bytes) -> str:
    """Extract text from Excel bytes using openpyxl."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(contents), data_only=True)
        text_parts = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                row_text = " ".join(str(cell) for cell in row if cell is not None)
                if row_text.strip():
                    text_parts.append(row_text)
        return "\n".join(text_parts)
    except Exception as e:
        print(f"Excel extraction error: {e}")
        return ""


def _extract_text(contents: bytes, filename: str) -> str:
    """Extract text from uploaded file based on extension."""
    ext = (filename or "").rsplit(".", 1)[-1].lower()
    if ext == "pdf":
        return _extract_text_from_pdf(contents)
    elif ext in ("xlsx", "xls"):
        return _extract_text_from_excel(contents)
    elif ext == "csv":
        try:
            return contents.decode("utf-8", errors="ignore")
        except:
            return ""
    return ""


def _extract_project_meta(text: str) -> dict:
    """Extract project-level metadata from document text."""
    import re
    meta = {}

    # Capacity patterns
    mw_match = re.search(r"(\d+)\s*MW\s*/\s*(\d+)\s*MWh", text)
    if mw_match:
        meta["capacity"] = f"{mw_match.group(1)} MW / {mw_match.group(2)} MWh"

    # Location
    locations = re.findall(r"(?:Tamil Nadu|Rajasthan|Bihar|Kerala|Maharashtra|Gujarat|Uttar Pradesh|Madhya Pradesh|Karnataka|Andhra Pradesh|Telangana|Haryana|Punjab|West Bengal|Odisha|Jharkhand|Chhattisgarh|Bikaner|Chennai|Barh|Patna|Sitapur|Kochi)", text, re.IGNORECASE)
    if locations:
        meta["location"] = f"{locations[0]}, India"

    # RFQ reference
    ref_match = re.search(r"(?:RFQ|IFB|Tender)\s*(?:Reference|No\.?|Number)[\s:]*([A-Z0-9\-/]+)", text)
    if ref_match:
        meta["rfq_ref"] = ref_match.group(1)

    # Solar PV
    solar_match = re.search(r"(\d+)\s*MWp", text)
    if solar_match:
        meta["solar_pv"] = f"{solar_match.group(1)} MWp"

    # Design life
    life_match = re.search(r"design\s*life.*?(\d+)\s*years", text, re.IGNORECASE)
    if life_match:
        meta["design_life"] = f"{life_match.group(1)} Years"

    # Issuer
    issuers = re.findall(r"(NTPC|Evolve Energy|BSPGCL|TNGECL|Eagle Infra|NHPC|SECI|PGCIL)", text, re.IGNORECASE)
    if issuers:
        meta["issuer"] = issuers[0].upper()

    return meta


# ─── Endpoints ───

@router.get("/")
async def list_rfqs():
    return {"items": RFQS, "total": len(RFQS), "page": 1, "per_page": 20}


@router.post("/upload")
async def upload_rfq(
    file: UploadFile = File(...),
    customer_name: str = Form(...),
    project_name: str = Form(...),
):
    """Upload any RFQ document (PDF/Excel/CSV) and extract technical requirements.

    The system:
    1. Reads the actual document text (PDF → PyPDF2, Excel → openpyxl)
    2. Scans for 50+ BESS technical parameters using keyword matching
    3. Extracts values (capacities, percentages, certifications, etc.)
    4. Returns structured requirements for OEM comparison
    """
    contents = await file.read()
    file_size = len(contents)
    file_ext = (file.filename or "").rsplit(".", 1)[-1].lower()

    # Step 1: Extract text from the document
    document_text = _extract_text(contents, file.filename or "")
    text_length = len(document_text)

    # Step 2: Extract technical requirements from the text
    from app.data.rfq_extraction import extract_from_text
    extracted_requirements = extract_from_text(document_text)

    # Step 3: Extract project metadata
    project_meta = _extract_project_meta(document_text)
    if not project_meta.get("issuer"):
        project_meta["issuer"] = customer_name

    # Step 4: Create and store the RFQ
    new_id = f"rfq-{len(RFQS) + 1:03d}"
    new_rfq = {
        "id": new_id,
        "customer_name": customer_name,
        "project_name": project_name,
        "status": "active",
        "created_at": "2026-03-19T10:00:00Z",
        "requirements": extracted_requirements,
        "source_file": file.filename,
        "file_size": file_size,
        "file_type": file_ext,
        "extraction_method": "document_parsing",
        "text_extracted": text_length > 0,
        "text_length": text_length,
    }
    if project_meta:
        new_rfq["project_meta"] = project_meta
    RFQS.append(new_rfq)

    return {
        "status": "extracted",
        "rfq_id": new_id,
        "customer_name": customer_name,
        "project_name": project_name,
        "file_name": file.filename,
        "file_size_bytes": file_size,
        "text_extracted": text_length > 0,
        "text_length": text_length,
        "requirements_extracted": len(extracted_requirements),
        "requirements": extracted_requirements,
        "project_meta": project_meta,
        "message": f"Extracted {len(extracted_requirements)} technical requirements from '{file.filename}' ({text_length} chars parsed)",
    }


@router.get("/{rfq_id}")
async def get_rfq(rfq_id: str):
    rfq = next((r for r in RFQS if r["id"] == rfq_id), None)
    if not rfq:
        return {"error": "RFQ not found"}
    return rfq


@router.get("/{rfq_id}/match")
async def match_rfq(rfq_id: str):
    rfq = next((r for r in RFQS if r["id"] == rfq_id), None)
    if not rfq:
        return {"error": "RFQ not found"}

    results = []
    for comp in COMPONENTS:
        params = PARAMETERS.get(comp["id"], [])
        if not params:
            continue

        param_map = {p["code"]: p for p in params}
        total_reqs = len(rfq["requirements"])
        passed = 0
        details = []

        for req in rfq["requirements"]:
            oem_param = param_map.get(req["code"])
            if not oem_param:
                details.append({
                    "parameter": req["parameter"],
                    "required": req["required_value"],
                    "oem_value": "N/A",
                    "match": False,
                    "section": req.get("section", ""),
                })
                continue

            oem_val = oem_param["value"]
            required = req["required_value"]
            match = False

            if required.startswith(">="):
                try:
                    match = float(oem_val) >= float(required[2:])
                except (ValueError, TypeError):
                    match = False
            elif required.startswith("<="):
                try:
                    match = float(oem_val) <= float(required[2:])
                except (ValueError, TypeError):
                    match = False
            elif required.lower() in ("yes", "true", "required"):
                match = oem_val.lower() in ("yes", "true", "certified", "pass")
            else:
                match = (
                    oem_val.strip().lower() == required.strip().lower()
                    or required.lower() in oem_val.lower()
                )

            if match:
                passed += 1

            details.append({
                "parameter": req["parameter"],
                "required": required,
                "oem_value": f"{oem_val} {oem_param['unit']}".strip(),
                "match": match,
                "section": req.get("section", ""),
            })

        match_pct = round((passed / total_reqs) * 100, 1) if total_reqs > 0 else 0
        results.append({
            "component_id": comp["id"],
            "model_name": comp["model_name"],
            "oem_name": comp["oem_name"],
            "match_percentage": match_pct,
            "passed": passed,
            "total": total_reqs,
            "details": details,
        })

    results.sort(key=lambda x: x["match_percentage"], reverse=True)
    return {"rfq_id": rfq_id, "customer": rfq["customer_name"], "matches": results}


class RFQCreate(BaseModel):
    customer_name: str
    project_name: str
    requirements: List[dict]


@router.post("/")
async def create_rfq(rfq: RFQCreate):
    new_id = f"rfq-{len(RFQS) + 1:03d}"
    new_rfq = {
        "id": new_id,
        "customer_name": rfq.customer_name,
        "project_name": rfq.project_name,
        "status": "active",
        "created_at": "2026-03-19T10:00:00Z",
        "requirements": rfq.requirements,
    }
    RFQS.append(new_rfq)
    return new_rfq
