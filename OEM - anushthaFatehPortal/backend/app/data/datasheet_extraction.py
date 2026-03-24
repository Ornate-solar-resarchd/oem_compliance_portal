"""
OEM Datasheet Extraction Engine
Extracts technical specs from uploaded datasheets (PDF/Excel) using Gemini AI.
Supports categories: Cell, DC Block, PCS, EMS
"""
import os
import json
import re
import io

CATEGORY_PROMPTS = {
    "Cell": """Extract ALL battery cell technical specifications:
- Nominal Capacity (Ah), Nominal Voltage (V), Energy (Wh/kWh)
- Internal Resistance (mΩ), Cycle Life, Calendar Life
- Max Charge/Discharge Rate (C-rate), Max Current (A)
- Weight (kg), Dimensions (L×W×H mm), Energy Density (Wh/kg)
- Operating Temperature range (charge/discharge/storage)
- Chemistry (LFP/NMC/etc), Form Factor (Prismatic/Cylindrical/Pouch/Blade)
- Certifications (IEC 62619, UL 1973, UN 38.3, BIS, UL 9540A)
- Self-discharge rate, Coulombic efficiency, Round-trip efficiency
- EOL capacity retention, SOC operating range
- Any other specs mentioned""",

    "DC Block": """Extract ALL DC Block / Battery Container / DC System specifications:
- Rated Energy (kWh/MWh), Rated Power (kW/MW)
- DC Voltage range (min/max/nominal)
- Number of battery modules/racks/clusters
- Cooling type (liquid/air), Cooling capacity
- Container dimensions (L×W×H), Weight
- Protection rating (IP rating), Operating temperature
- BMS type, Communication protocol
- Fire suppression system type
- Certifications, Safety features
- Any other specs mentioned""",

    "PCS": """Extract ALL Power Conversion System (PCS) / Inverter specifications:
- Rated AC Power (kW/MW), Rated DC Power
- AC Voltage (V), DC Voltage range
- Frequency (Hz), Power Factor range
- Max Efficiency, Euro/CEC efficiency
- THD (Total Harmonic Distortion)
- Cooling type, Operating temperature
- Dimensions, Weight
- Communication protocols (Modbus/CAN/Ethernet)
- Grid support features (PFR, SFR, Black Start, Reactive Power)
- Certifications (IEC 62477, IEC 61000, IEEE 1547)
- Protection features
- Any other specs mentioned""",

    "EMS": """Extract ALL Energy Management System (EMS) / SCADA specifications:
- System architecture (server config, redundancy)
- Data refresh rate, System availability (%)
- Communication protocols (IEC 61850, Modbus TCP/IP, DNP3)
- Control modes (peak shaving, frequency regulation, load shifting, etc.)
- SCADA integration, HMI details
- Cyber security features (RBAC, firewall, audit trail)
- GPS time synchronization
- Remote monitoring capabilities
- Software origin (indigenous/imported)
- Hardware requirements (servers, RTUs, gateways)
- Certifications (CEA, IEC standards)
- Any other specs mentioned""",
}


def extract_text_from_file(contents: bytes, filename: str) -> str:
    """Extract text from PDF or Excel file."""
    ext = (filename or "").rsplit(".", 1)[-1].lower()

    if ext == "pdf":
        try:
            from PyPDF2 import PdfReader
            reader = PdfReader(io.BytesIO(contents))
            return "\n".join(p.extract_text() or "" for p in reader.pages)
        except Exception as e:
            print(f"PDF extraction error: {e}")
            return ""

    elif ext in ("xlsx", "xls"):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(contents), data_only=True)
            lines = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    row_text = " ".join(str(c) for c in row if c is not None)
                    if row_text.strip():
                        lines.append(row_text)
            return "\n".join(lines)
        except Exception as e:
            print(f"Excel extraction error: {e}")
            return ""

    elif ext == "csv":
        return contents.decode("utf-8", errors="ignore")

    return ""


def extract_specs_with_gemini(text: str, category: str, api_key: str) -> list:
    """Use Gemini AI to extract specs from datasheet text."""
    try:
        import google.generativeai as genai
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel("gemini-2.0-flash")

        category_prompt = CATEGORY_PROMPTS.get(category, CATEGORY_PROMPTS["Cell"])

        prompt = f"""You are an expert BESS (Battery Energy Storage System) technical engineer.
Your job: Extract EVERY SINGLE technical specification, parameter, value, rating, and measurement from this {category} datasheet.

BE EXHAUSTIVE. Extract EVERYTHING — even minor specs. The goal is to capture 30-50+ parameters minimum.

{category_prompt}

IMPORTANT RULES:
1. Extract EVERY number, rating, spec, measurement, range, limit, threshold mentioned in the document
2. If a parameter has a range (e.g., "0°C to 55°C"), extract both min and max as separate entries
3. Include ALL certifications, standards, and compliance marks individually
4. Include model number, manufacturer, part number, dimensions, weight
5. Include operating conditions, storage conditions, transportation conditions
6. Include warranty, design life, degradation specs
7. Include communication protocols, connector types, mounting details
8. If something is mentioned but no exact value, extract it with "See datasheet" as value
9. DO NOT skip any specification — more is better

DATASHEET TEXT:
{text[:40000]}

Return a JSON array of objects, each with:
- "name": parameter name (human readable, descriptive)
- "code": short uppercase code (e.g. CELL_CAPACITY_AH, PCS_RATED_POWER, DC_COOLING_TYPE)
- "value": the exact extracted value as string
- "unit": unit of measurement (V, A, Ah, mm, kg, °C, %, W, kW, MW, MWh, etc.)
- "section": category (Electrical, Physical, Thermal, Safety, Performance, Communication, Mechanical, General)
- "status": "pass"

CRITICAL: Return AT LEAST 20 parameters. Extract every single data point. Return ONLY the JSON array. Start with [ end with ]."""

        response = model.generate_content(prompt)
        response_text = response.text.strip()

        if response_text.startswith("```"):
            response_text = re.sub(r'^```(?:json)?\s*', '', response_text)
            response_text = re.sub(r'\s*```$', '', response_text)

        if response_text.startswith("["):
            return json.loads(response_text)

        match = re.search(r'\[.*\]', response_text, re.DOTALL)
        if match:
            return json.loads(match.group())

    except Exception as e:
        print(f"Gemini datasheet extraction error: {e}")

    return []


def extract_specs_keyword(text: str, category: str) -> list:
    """Fallback keyword-based extraction when no AI available."""
    text_lower = text.lower()
    extracted = []

    if category == "Cell":
        patterns = [
            # Electrical
            ("Nominal Capacity", "CELL_CAPACITY_AH", r"(?:nominal|rated|typical)\s*capacity.*?(\d+(?:\.\d+)?)\s*Ah", "Ah", "Electrical"),
            ("Minimum Capacity", "CELL_MIN_CAPACITY_AH", r"(?:minimum|min)\s*capacity.*?(\d+(?:\.\d+)?)\s*Ah", "Ah", "Electrical"),
            ("Nominal Voltage", "CELL_VOLTAGE_V", r"(?:nominal|rated)\s*voltage.*?(\d+(?:\.\d+)?)\s*V", "V", "Electrical"),
            ("Charge Cut-off Voltage", "CELL_CHARGE_CUTOFF_V", r"(?:charge|charging)\s*(?:cut[\s-]*off|upper)\s*voltage.*?(\d+(?:\.\d+)?)\s*V", "V", "Electrical"),
            ("Discharge Cut-off Voltage", "CELL_DISCHARGE_CUTOFF_V", r"(?:discharge|discharging)\s*(?:cut[\s-]*off|lower)\s*voltage.*?(\d+(?:\.\d+)?)\s*V", "V", "Electrical"),
            ("Energy", "CELL_ENERGY_WH", r"energy.*?(\d+(?:\.\d+)?)\s*(?:Wh|kWh)", "Wh", "Electrical"),
            ("Internal Resistance", "CELL_IR_MOHM", r"(?:internal\s*)?(?:AC\s*)?resistance.*?(\d+(?:\.\d+)?)\s*m[Ωo]", "mΩ", "Electrical"),
            ("Max Charge Current", "CELL_MAX_CHARGE_A", r"(?:max|maximum)\s*(?:charge|charging)\s*current.*?(\d+(?:\.\d+)?)\s*A", "A", "Electrical"),
            ("Max Discharge Current", "CELL_MAX_DISCHARGE_A", r"(?:max|maximum)\s*(?:discharge|discharging)\s*current.*?(\d+(?:\.\d+)?)\s*A", "A", "Electrical"),
            ("Standard Charge Current", "CELL_STD_CHARGE_A", r"(?:standard|recommended)\s*(?:charge|charging)\s*current.*?(\d+(?:\.\d+)?)\s*A", "A", "Electrical"),
            ("Standard Discharge Current", "CELL_STD_DISCHARGE_A", r"(?:standard|recommended)\s*(?:discharge|discharging)\s*current.*?(\d+(?:\.\d+)?)\s*A", "A", "Electrical"),
            ("Max Charge Rate", "CELL_MAX_CHARGE_RATE", r"(?:max|maximum)\s*(?:charge|charging)\s*rate.*?(\d+(?:\.\d+)?)\s*C", "C", "Electrical"),
            ("Max Discharge Rate", "CELL_MAX_DISCHARGE_RATE", r"(?:max|maximum)\s*(?:discharge|discharging)\s*rate.*?(\d+(?:\.\d+)?)\s*C", "C", "Electrical"),
            # Life
            ("Cycle Life", "CELL_CYCLE_LIFE", r"cycle\s*life.*?(\d[\d,]*)\s*(?:cycles|times)", "cycles", "Performance"),
            ("Calendar Life", "CELL_CALENDAR_LIFE", r"calendar\s*life.*?(\d+)\s*years?", "years", "Performance"),
            ("EOL Capacity Retention", "CELL_EOL_RETENTION", r"(?:EOL|end\s*of\s*life|capacity\s*retention).*?(\d+(?:\.\d+)?)\s*%", "%", "Performance"),
            ("Self-Discharge Rate", "CELL_SELF_DISCHARGE", r"self[\s-]*discharge.*?(\d+(?:\.\d+)?)\s*%", "%/month", "Performance"),
            ("Coulombic Efficiency", "CELL_COULOMBIC_EFF", r"coulombic\s*efficiency.*?(\d+(?:\.\d+)?)\s*%", "%", "Performance"),
            ("Round-trip Efficiency", "CELL_RTE", r"(?:round[\s-]*trip|energy)\s*efficiency.*?(\d+(?:\.\d+)?)\s*%", "%", "Performance"),
            # Physical
            ("Weight", "CELL_WEIGHT_KG", r"weight.*?(\d+(?:\.\d+)?)\s*(?:kg|g)", "kg", "Physical"),
            ("Length", "CELL_LENGTH_MM", r"length.*?(\d+(?:\.\d+)?)\s*mm", "mm", "Physical"),
            ("Width", "CELL_WIDTH_MM", r"width.*?(\d+(?:\.\d+)?)\s*mm", "mm", "Physical"),
            ("Height / Thickness", "CELL_HEIGHT_MM", r"(?:height|thickness).*?(\d+(?:\.\d+)?)\s*mm", "mm", "Physical"),
            ("Energy Density (Gravimetric)", "CELL_ENERGY_DENSITY_WH_KG", r"(?:energy|gravimetric)\s*density.*?(\d+(?:\.\d+)?)\s*Wh/kg", "Wh/kg", "Physical"),
            ("Energy Density (Volumetric)", "CELL_ENERGY_DENSITY_WH_L", r"(?:volumetric)\s*.*?density.*?(\d+(?:\.\d+)?)\s*Wh/[Ll]", "Wh/L", "Physical"),
            # Thermal
            ("Charge Temp Min", "CELL_CHARGE_TEMP_MIN", r"(?:charge|charging)\s*(?:temp|temperature).*?(-?\d+)\s*[°℃C]", "°C", "Thermal"),
            ("Charge Temp Max", "CELL_CHARGE_TEMP_MAX", r"(?:charge|charging)\s*(?:temp|temperature).*?(\d+)\s*[°℃C]", "°C", "Thermal"),
            ("Discharge Temp Min", "CELL_DISCHARGE_TEMP_MIN", r"(?:discharge|discharging)\s*(?:temp|temperature).*?(-?\d+)\s*[°℃C]", "°C", "Thermal"),
            ("Discharge Temp Max", "CELL_DISCHARGE_TEMP_MAX", r"(?:discharge|discharging)\s*(?:temp|temperature).*?(\d+)\s*[°℃C]", "°C", "Thermal"),
            ("Storage Temp Min", "CELL_STORAGE_TEMP_MIN", r"storage\s*(?:temp|temperature).*?(-?\d+)\s*[°℃C]", "°C", "Thermal"),
            ("Storage Temp Max", "CELL_STORAGE_TEMP_MAX", r"storage\s*(?:temp|temperature).*?(\d+)\s*[°℃C]", "°C", "Thermal"),
            # General
            ("SOC Range", "CELL_SOC_RANGE", r"SOC.*?(\d+)\s*%?\s*[-~to]\s*(\d+)\s*%", "%", "Electrical"),
        ]
    elif category == "DC Block":
        patterns = [
            ("Rated Energy", "DC_RATED_ENERGY", r"(?:rated|nominal)\s*energy.*?(\d+(?:\.\d+)?)\s*(?:kWh|MWh)", "kWh", "System"),
            ("Rated Power", "DC_RATED_POWER", r"(?:rated|nominal)\s*power.*?(\d+(?:\.\d+)?)\s*(?:kW|MW)", "kW", "System"),
            ("Usable Energy", "DC_USABLE_ENERGY", r"usable\s*energy.*?(\d+(?:\.\d+)?)\s*(?:kWh|MWh)", "kWh", "System"),
            ("DC Voltage Nominal", "DC_VOLTAGE_NOM", r"(?:dc|battery|nominal)\s*voltage.*?(\d+(?:\.\d+)?)\s*V", "V", "Electrical"),
            ("DC Voltage Range Max", "DC_VOLTAGE_MAX", r"(?:max|maximum)\s*(?:dc\s*)?voltage.*?(\d+(?:\.\d+)?)\s*V", "V", "Electrical"),
            ("DC Voltage Range Min", "DC_VOLTAGE_MIN", r"(?:min|minimum)\s*(?:dc\s*)?voltage.*?(\d+(?:\.\d+)?)\s*V", "V", "Electrical"),
            ("Number of Modules", "DC_MODULE_COUNT", r"(\d+)\s*(?:modules|racks|battery\s*packs)", "", "System"),
            ("Number of Clusters", "DC_CLUSTER_COUNT", r"(\d+)\s*clusters", "", "System"),
            ("Cooling Type", "DC_COOLING", r"(liquid|air)\s*cool", "", "Thermal"),
            ("Cooling Capacity", "DC_COOLING_KW", r"cool.*?capacity.*?(\d+(?:\.\d+)?)\s*kW", "kW", "Thermal"),
            ("Container Length", "DC_LENGTH", r"(?:container\s*)?length.*?(\d+(?:\.\d+)?)\s*(?:mm|m|ft)", "mm", "Physical"),
            ("Container Width", "DC_WIDTH", r"(?:container\s*)?width.*?(\d+(?:\.\d+)?)\s*(?:mm|m|ft)", "mm", "Physical"),
            ("Container Height", "DC_HEIGHT", r"(?:container\s*)?height.*?(\d+(?:\.\d+)?)\s*(?:mm|m|ft)", "mm", "Physical"),
            ("Total Weight", "DC_WEIGHT", r"(?:total\s*)?weight.*?(\d+(?:\.\d+)?)\s*(?:kg|ton)", "kg", "Physical"),
            ("Protection Rating", "DC_IP_RATING", r"(IP\d+\w?)", "", "Safety"),
            ("Operating Temp Min", "DC_OP_TEMP_MIN", r"operating\s*temp.*?(-?\d+)\s*[°℃C]", "°C", "Thermal"),
            ("Operating Temp Max", "DC_OP_TEMP_MAX", r"operating\s*temp.*?(\d+)\s*[°℃C]", "°C", "Thermal"),
            ("Fire Suppression", "DC_FIRE_SYSTEM", r"(aerosol|gas|water\s*mist|NOVEC|FM-?200)", "", "Safety"),
            ("BMS Type", "DC_BMS", r"BMS.*?(\w[\w\s]*)", "", "Communication"),
            ("Communication", "DC_COMM", r"(CAN|RS485|Modbus|Ethernet|TCP/IP)", "", "Communication"),
            ("RTE", "DC_RTE", r"(?:round[\s-]*trip|RTE).*?(\d+(?:\.\d+)?)\s*%", "%", "Performance"),
        ]
    elif category == "PCS":
        patterns = [
            ("Rated AC Power", "PCS_AC_POWER", r"(?:rated|nominal)\s*(?:ac\s*)?power.*?(\d+(?:\.\d+)?)\s*(?:kW|MW)", "kW", "Electrical"),
            ("Max AC Power", "PCS_MAX_AC_POWER", r"(?:max|maximum)\s*(?:ac\s*)?power.*?(\d+(?:\.\d+)?)\s*(?:kW|MW)", "kW", "Electrical"),
            ("Rated DC Power", "PCS_DC_POWER", r"(?:rated|nominal)\s*dc\s*power.*?(\d+(?:\.\d+)?)\s*(?:kW|MW)", "kW", "Electrical"),
            ("AC Voltage", "PCS_AC_VOLTAGE", r"(?:ac\s*)?(?:output\s*)?voltage.*?(\d+(?:\.\d+)?)\s*V", "V", "Electrical"),
            ("DC Voltage Min", "PCS_DC_VOLTAGE_MIN", r"(?:dc\s*)?(?:input\s*)?(?:voltage\s*)?(?:min|minimum).*?(\d+(?:\.\d+)?)\s*V", "V", "Electrical"),
            ("DC Voltage Max", "PCS_DC_VOLTAGE_MAX", r"(?:dc\s*)?(?:input\s*)?(?:voltage\s*)?(?:max|maximum).*?(\d+(?:\.\d+)?)\s*V", "V", "Electrical"),
            ("Max Efficiency", "PCS_EFFICIENCY", r"(?:max|peak|maximum)\s*efficiency.*?(\d+(?:\.\d+)?)\s*%", "%", "Performance"),
            ("Euro Efficiency", "PCS_EURO_EFF", r"(?:euro|european|CEC)\s*efficiency.*?(\d+(?:\.\d+)?)\s*%", "%", "Performance"),
            ("THD", "PCS_THD", r"THD.*?<?\s*(\d+(?:\.\d+)?)\s*%", "%", "Performance"),
            ("Power Factor", "PCS_POWER_FACTOR", r"power\s*factor.*?(\d+(?:\.\d+)?)", "", "Electrical"),
            ("Frequency", "PCS_FREQUENCY", r"(\d+(?:\.\d+)?)\s*Hz", "Hz", "Electrical"),
            ("Max AC Current", "PCS_MAX_AC_CURRENT", r"(?:max|maximum)\s*(?:ac\s*)?(?:output\s*)?current.*?(\d+(?:\.\d+)?)\s*A", "A", "Electrical"),
            ("Cooling Type", "PCS_COOLING", r"(liquid|air|forced\s*air)\s*cool", "", "Thermal"),
            ("Operating Temp Min", "PCS_OP_TEMP_MIN", r"operating\s*temp.*?(-?\d+)\s*[°℃C]", "°C", "Thermal"),
            ("Operating Temp Max", "PCS_OP_TEMP_MAX", r"operating\s*temp.*?(\d+)\s*[°℃C]", "°C", "Thermal"),
            ("Dimensions L", "PCS_LENGTH", r"(?:dimension|size).*?(\d+)\s*(?:×|x|X)\s*\d+\s*(?:×|x|X)\s*\d+\s*mm", "mm", "Physical"),
            ("Weight", "PCS_WEIGHT", r"weight.*?(\d+(?:\.\d+)?)\s*(?:kg|ton)", "kg", "Physical"),
            ("Protection Rating", "PCS_IP_RATING", r"(IP\d+\w?)", "", "Safety"),
            ("Communication", "PCS_COMM", r"(Modbus|CAN|Ethernet|RS485|TCP/IP|IEC 61850)", "", "Communication"),
            ("Grid Support - PFR", "PCS_PFR", r"(?:primary\s*frequency|PFR)", "", "Grid Support"),
            ("Grid Support - Reactive", "PCS_REACTIVE", r"reactive\s*power", "", "Grid Support"),
            ("Grid Support - Black Start", "PCS_BLACK_START", r"black\s*start", "", "Grid Support"),
            ("Topology", "PCS_TOPOLOGY", r"(H-?bridge|NPC|T-?type|3[\s-]*level)", "", "Electrical"),
        ]
    elif category == "EMS":
        patterns = [
            ("System Availability", "EMS_AVAILABILITY", r"availability.*?(\d+(?:\.\d+)?)\s*%", "%", "Performance"),
            ("Data Refresh Rate", "EMS_REFRESH", r"(?:refresh|scan|update).*?(?:≤|<=)?\s*(\d+)\s*(?:s|sec|second)", "s", "Communication"),
            ("Communication Protocol", "EMS_PROTOCOL", r"(IEC 61850|Modbus TCP/IP|Modbus|DNP3)", "", "Communication"),
            ("Server Configuration", "EMS_SERVER", r"(redundant|dual|single)\s*server", "", "Architecture"),
            ("Time Sync", "EMS_TIME_SYNC", r"(GPS|NTP|PTP)\s*(?:based|time|sync)", "", "Communication"),
            ("Cyber Security", "EMS_CYBERSEC", r"(RBAC|role[\s-]*based|firewall|CEA\s*cyber)", "", "Security"),
            ("SCADA Integration", "EMS_SCADA", r"SCADA\s*(?:integration|interface)", "", "Communication"),
            ("Control Modes", "EMS_MODES", r"(peak\s*shaving|frequency\s*regulation|load\s*shifting|arbitrage|ramp\s*rate)", "", "Functionality"),
            ("Software Origin", "EMS_ORIGIN", r"(indigenous|Make in India|developed in India)", "", "Compliance"),
            ("RTE Monitoring", "EMS_RTE_MON", r"RTE\s*(?:monitoring|measurement|tracking)", "", "Functionality"),
            ("Black Start Logic", "EMS_BLACK_START", r"black\s*start\s*(?:logic|capability|support)", "", "Functionality"),
            ("Data Historian", "EMS_HISTORIAN", r"(data\s*historian|historical\s*data|data\s*logging)", "", "Architecture"),
            ("HMI", "EMS_HMI", r"(HMI|human[\s-]*machine[\s-]*interface|workstation)", "", "Architecture"),
        ]
    else:
        patterns = []

    for name, code, pattern, unit, section in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            extracted.append({
                "name": name, "code": code, "value": match.group(1).replace(",", ""),
                "unit": unit, "section": section, "status": "pass", "confidence": 0.85,
            })

    # Check certifications (all categories)
    certs = []
    cert_checks = [
        ("iec 62619", "IEC 62619"), ("ul 1973", "UL 1973"), ("ul 9540a", "UL 9540A"),
        ("ul 9540", "UL 9540"), ("un 38.3", "UN 38.3"), ("un38.3", "UN 38.3"),
        ("iec 62933", "IEC 62933"), ("iec 62477", "IEC 62477"), ("iec 61000", "IEC 61000"),
        ("iec 61850", "IEC 61850"), ("ieee 1547", "IEEE 1547"), ("nfpa", "NFPA"),
        ("bis", "BIS"), ("cea", "CEA"), ("cerc", "CERC"),
    ]
    for keyword, cert_name in cert_checks:
        if keyword in text_lower:
            certs.append(cert_name)
    if certs:
        extracted.append({"name": "Certifications", "code": f"{category.upper().replace(' ', '_')}_CERTS",
                         "value": ", ".join(certs), "unit": "", "section": "Safety", "status": "pass", "confidence": 0.95})

    # Chemistry / Technology (for all categories)
    chemistry_map = {"lfp": "LFP", "lifepo4": "LFP", "nmc": "NMC", "lto": "LTO", "sodium": "Sodium-ion"}
    for kw, val in chemistry_map.items():
        if kw in text_lower:
            extracted.append({"name": "Chemistry", "code": f"{category.upper().replace(' ', '_')}_CHEMISTRY",
                             "value": val, "unit": "", "section": "General", "status": "pass", "confidence": 0.92})
            break

    # Form Factor
    form_factors = {"prismatic": "Prismatic", "cylindrical": "Cylindrical", "pouch": "Pouch", "blade": "Blade"}
    for kw, val in form_factors.items():
        if kw in text_lower:
            extracted.append({"name": "Form Factor", "code": f"{category.upper().replace(' ', '_')}_FORM",
                             "value": val, "unit": "", "section": "Physical", "status": "pass", "confidence": 0.90})
            break

    # Model Number — try to find it
    model_match = re.search(r"(?:model|type|part)\s*(?:no|number|#)?[\s:]*([A-Z0-9][\w\-]{3,})", text, re.IGNORECASE)
    if model_match:
        extracted.append({"name": "Model Number", "code": f"{category.upper().replace(' ', '_')}_MODEL",
                         "value": model_match.group(1), "unit": "", "section": "General", "status": "pass", "confidence": 0.88})

    # Manufacturer
    mfr_match = re.search(r"(?:manufacturer|brand|made\s*by|company)[\s:]*([A-Z][\w\s]{2,30})", text, re.IGNORECASE)
    if mfr_match:
        extracted.append({"name": "Manufacturer", "code": f"{category.upper().replace(' ', '_')}_MFR",
                         "value": mfr_match.group(1).strip(), "unit": "", "section": "General", "status": "pass", "confidence": 0.85})

    # Design Life
    life_match = re.search(r"(?:design|service|expected)\s*life.*?(\d+)\s*years?", text, re.IGNORECASE)
    if life_match:
        extracted.append({"name": "Design Life", "code": f"{category.upper().replace(' ', '_')}_DESIGN_LIFE",
                         "value": life_match.group(1), "unit": "years", "section": "Performance", "status": "pass", "confidence": 0.90})

    # Warranty
    warranty_match = re.search(r"warranty.*?(\d+)\s*years?", text, re.IGNORECASE)
    if warranty_match:
        extracted.append({"name": "Warranty Period", "code": f"{category.upper().replace(' ', '_')}_WARRANTY",
                         "value": warranty_match.group(1), "unit": "years", "section": "General", "status": "pass", "confidence": 0.88})

    return extracted


def extract_from_datasheet(contents: bytes, filename: str, category: str) -> list:
    """Main extraction function — tries Gemini first, falls back to keywords."""
    text = extract_text_from_file(contents, filename)
    if not text:
        return []

    gemini_key = os.environ.get("GEMINI_API_KEY", "")
    if gemini_key and gemini_key != "YOUR_GEMINI_KEY_HERE" and len(gemini_key) > 10:
        print(f"[Datasheet] Using Gemini AI for {category} extraction...")
        results = extract_specs_with_gemini(text, category, gemini_key)
        if results:
            print(f"[Datasheet] Gemini extracted {len(results)} specs")
            return results

    print(f"[Datasheet] Using keyword extraction for {category}...")
    return extract_specs_keyword(text, category)
