"""
Celery task definitions for async operations.

Tasks:
  1. generate_document  — render compliance PDFs / Excel exports, upload to MinIO
  2. run_ai_extraction  — extract parameter values from datasheets via Claude API
"""
from __future__ import annotations

import hashlib
import io
import json
import logging
import re
from datetime import datetime, timezone
from uuid import UUID

import boto3
from botocore.client import Config as BotoConfig
from celery import Celery
from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session, sessionmaker

from app.core.config import settings

# ---------------------------------------------------------------------------
# Celery app
# ---------------------------------------------------------------------------
celery_app = Celery("tcp", broker=settings.REDIS_URL, backend=settings.REDIS_URL)
celery_app.config_from_object("app.workers.celeryconfig")

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Sync database session (Celery workers cannot use asyncio)
# ---------------------------------------------------------------------------
_sync_url = settings.DATABASE_URL_SYNC or settings.DATABASE_URL.replace(
    "postgresql+asyncpg://", "postgresql://"
).replace(
    "postgresql+aiopg://", "postgresql://"
)

sync_engine = create_engine(
    _sync_url,
    pool_size=5,
    max_overflow=10,
    pool_pre_ping=True,
    echo=settings.APP_ENV == "development",
)

SyncSessionLocal = sessionmaker(bind=sync_engine, expire_on_commit=False)


def get_sync_session() -> Session:
    """Return a new synchronous SQLAlchemy session."""
    return SyncSessionLocal()


# ---------------------------------------------------------------------------
# MinIO / S3 helper
# ---------------------------------------------------------------------------
def _get_s3_client():
    """Return a boto3 S3 client pointed at the MinIO endpoint."""
    scheme = "https" if settings.MINIO_SECURE else "http"
    return boto3.client(
        "s3",
        endpoint_url=f"{scheme}://{settings.MINIO_ENDPOINT}",
        aws_access_key_id=settings.MINIO_ACCESS_KEY,
        aws_secret_access_key=settings.MINIO_SECRET_KEY,
        config=BotoConfig(signature_version="s3v4"),
        region_name="us-east-1",
    )


def _ensure_bucket(s3, bucket_name: str) -> None:
    """Create the bucket if it does not already exist."""
    try:
        s3.head_bucket(Bucket=bucket_name)
    except Exception:
        try:
            s3.create_bucket(Bucket=bucket_name)
        except Exception:
            pass  # may already exist in a race


# ---------------------------------------------------------------------------
# Lazy model imports (avoid circular imports at module level)
# ---------------------------------------------------------------------------
def _import_models():
    """Import ORM models lazily so Celery can discover tasks without
    triggering the full FastAPI import tree at module level."""
    from app.models.document import Document, DocumentFormat, DocumentType
    from app.models.ai_extraction import AIExtractionJob
    from app.models.oem import ComponentModel, ComponentType
    from app.models.parameter import StandardParameter, ParameterValue, ExtractionSource
    from app.models.sheet import TechnicalSheet, SheetComplianceResult
    from app.models.project import Project

    return {
        "Document": Document,
        "DocumentFormat": DocumentFormat,
        "DocumentType": DocumentType,
        "AIExtractionJob": AIExtractionJob,
        "ComponentModel": ComponentModel,
        "ComponentType": ComponentType,
        "StandardParameter": StandardParameter,
        "ParameterValue": ParameterValue,
        "ExtractionSource": ExtractionSource,
        "TechnicalSheet": TechnicalSheet,
        "SheetComplianceResult": SheetComplianceResult,
        "Project": Project,
    }


# ============================================================================
# TASK 1: generate_document
# ============================================================================
@celery_app.task(name="generate_document", bind=True, max_retries=3)
def generate_document(self, document_id: str):
    """
    Generate a compliance document (PDF or Excel), upload to MinIO,
    and mark the document row as ready.

    Workflow:
      1. Load the Document row and its related sheet / project data.
      2. Render content based on document_type and format.
      3. Upload the rendered file to MinIO.
      4. Update the document row with minio_key, file_size_bytes, sha256,
         is_ready=True, and ready_at.
    """
    models = _import_models()
    Document = models["Document"]
    TechnicalSheet = models["TechnicalSheet"]
    SheetComplianceResult = models["SheetComplianceResult"]
    ComponentModel = models["ComponentModel"]
    Project = models["Project"]

    session: Session = get_sync_session()
    try:
        # ------------------------------------------------------------------
        # 1. Load document
        # ------------------------------------------------------------------
        doc = session.get(Document, UUID(document_id))
        if doc is None:
            raise ValueError(f"Document {document_id} not found")

        doc_type = doc.document_type.value   # e.g. "compliance_sheet"
        doc_format = doc.format.value        # e.g. "pdf" or "excel"

        # Gather context data
        sheet = None
        project = None
        compliance_results = []
        component_model = None

        if doc.technical_sheet_id:
            sheet = session.get(TechnicalSheet, doc.technical_sheet_id)
            if sheet:
                compliance_results = (
                    session.execute(
                        select(SheetComplianceResult)
                        .where(SheetComplianceResult.technical_sheet_id == sheet.id)
                        .order_by(SheetComplianceResult.created_at)
                    )
                    .scalars()
                    .all()
                )
                component_model = session.get(ComponentModel, sheet.component_model_id)

        if doc.project_id:
            project = session.get(Project, doc.project_id)

        # ------------------------------------------------------------------
        # 2. Render content
        # ------------------------------------------------------------------
        if doc_format == "pdf":
            file_bytes, content_type = _render_pdf(
                doc_type, sheet, project, compliance_results, component_model
            )
            extension = "pdf"
        elif doc_format == "excel":
            file_bytes, content_type = _render_excel(
                doc_type, sheet, project, compliance_results, component_model
            )
            extension = "xlsx"
        else:
            # pptx — placeholder
            file_bytes = b""
            content_type = "application/vnd.openxmlformats-officedocument.presentationml.presentation"
            extension = "pptx"

        # ------------------------------------------------------------------
        # 3. Upload to MinIO
        # ------------------------------------------------------------------
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        safe_type = doc_type.replace(" ", "_")
        minio_key = f"documents/{document_id}/{safe_type}_{timestamp}.{extension}"

        s3 = _get_s3_client()
        bucket = settings.MINIO_BUCKET_DOCUMENTS
        _ensure_bucket(s3, bucket)

        s3.put_object(
            Bucket=bucket,
            Key=minio_key,
            Body=file_bytes,
            ContentType=content_type,
        )

        # ------------------------------------------------------------------
        # 4. Update document row
        # ------------------------------------------------------------------
        sha = hashlib.sha256(file_bytes).hexdigest()

        doc.minio_key = minio_key
        doc.file_size_bytes = len(file_bytes)
        doc.sha256 = sha
        doc.is_ready = True
        doc.ready_at = datetime.now(timezone.utc)
        doc.error_message = None

        session.commit()
        logger.info("Document %s generated successfully (%d bytes)", document_id, len(file_bytes))

    except Exception as exc:
        session.rollback()
        # Try to record the error on the document row
        try:
            doc = session.get(Document, UUID(document_id))
            if doc:
                doc.error_message = f"{type(exc).__name__}: {exc}"
                session.commit()
        except Exception:
            session.rollback()

        logger.exception("generate_document failed for %s", document_id)
        raise self.retry(exc=exc, countdown=30 * (self.request.retries + 1))
    finally:
        session.close()


# ---------------------------------------------------------------------------
# PDF rendering
# ---------------------------------------------------------------------------
def _render_pdf(doc_type, sheet, project, compliance_results, component_model):
    """Render an HTML template to PDF via WeasyPrint."""
    from weasyprint import HTML

    title = _doc_title(doc_type, sheet, project)
    rows_html = _build_results_table_html(compliance_results)

    sheet_info = ""
    if sheet:
        sheet_info = f"""
        <tr><td><strong>Sheet Number</strong></td><td>{sheet.sheet_number or 'N/A'}</td></tr>
        <tr><td><strong>Stage</strong></td><td>{sheet.stage.value if hasattr(sheet.stage, 'value') else sheet.stage}</td></tr>
        <tr><td><strong>Revision</strong></td><td>{sheet.revision}</td></tr>
        <tr><td><strong>Compliance Score</strong></td><td>{sheet.compliance_score or 'N/A'}%</td></tr>
        """

    model_info = ""
    if component_model:
        model_info = f"""
        <tr><td><strong>Component Model</strong></td><td>{component_model.model_name}</td></tr>
        """

    project_info = ""
    if project:
        project_info = f"""
        <tr><td><strong>Project</strong></td><td>{project.name}</td></tr>
        <tr><td><strong>Client</strong></td><td>{project.client_name}</td></tr>
        """

    html_content = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <style>
        body {{ font-family: 'Helvetica Neue', Arial, sans-serif; margin: 40px; color: #1a1a1a; font-size: 11pt; }}
        h1 {{ color: #1e3a5f; border-bottom: 3px solid #1e3a5f; padding-bottom: 10px; font-size: 18pt; }}
        h2 {{ color: #2c5282; margin-top: 30px; font-size: 14pt; }}
        table {{ border-collapse: collapse; width: 100%; margin-top: 15px; }}
        th, td {{ border: 1px solid #cbd5e0; padding: 8px 12px; text-align: left; font-size: 10pt; }}
        th {{ background-color: #1e3a5f; color: white; font-weight: 600; }}
        tr:nth-child(even) {{ background-color: #f7fafc; }}
        .info-table td {{ border: none; padding: 4px 12px; }}
        .info-table {{ margin-bottom: 20px; }}
        .pass {{ color: #276749; font-weight: bold; }}
        .fail {{ color: #c53030; font-weight: bold; }}
        .pending {{ color: #b7791f; }}
        .waived {{ color: #2b6cb0; }}
        .footer {{ margin-top: 40px; font-size: 9pt; color: #718096; border-top: 1px solid #e2e8f0; padding-top: 10px; }}
        .header-logo {{ font-size: 10pt; color: #718096; margin-bottom: 5px; }}
    </style>
</head>
<body>
    <div class="header-logo">UnityESS Technical Compliance Portal — Ornate Agencies Pvt. Ltd.</div>
    <h1>{title}</h1>

    <h2>Document Information</h2>
    <table class="info-table">
        {project_info}
        {sheet_info}
        {model_info}
        <tr><td><strong>Generated</strong></td><td>{datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}</td></tr>
    </table>

    {rows_html}

    <div class="footer">
        This document was auto-generated by the UnityESS Technical Compliance Portal.
        Content integrity can be verified via SHA-256 hash stored in the system.
    </div>
</body>
</html>"""

    pdf_bytes = HTML(string=html_content).write_pdf()
    return pdf_bytes, "application/pdf"


def _build_results_table_html(compliance_results) -> str:
    """Build an HTML table of compliance results."""
    if not compliance_results:
        return "<p><em>No compliance results available.</em></p>"

    rows = ""
    for r in compliance_results:
        status_val = r.status.value if hasattr(r.status, "value") else str(r.status)
        # Handle the pass_ enum name
        if status_val == "pass" or (hasattr(r.status, "name") and r.status.name == "pass_"):
            status_display = "PASS"
            css_class = "pass"
        elif status_val == "fail":
            status_display = "FAIL"
            css_class = "fail"
        elif status_val == "waived":
            status_display = "WAIVED"
            css_class = "waived"
        else:
            status_display = "PENDING"
            css_class = "pending"

        param_id_short = str(r.standard_parameter_id)[:8]
        rows += f"""<tr>
            <td>{param_id_short}...</td>
            <td>{r.oem_value or 'N/A'}</td>
            <td class="{css_class}">{status_display}</td>
            <td>{r.waive_reason or ''}</td>
        </tr>"""

    return f"""
    <h2>Compliance Results</h2>
    <table>
        <thead>
            <tr>
                <th>Parameter</th>
                <th>OEM Value</th>
                <th>Status</th>
                <th>Notes</th>
            </tr>
        </thead>
        <tbody>
            {rows}
        </tbody>
    </table>
    """


# ---------------------------------------------------------------------------
# Excel rendering
# ---------------------------------------------------------------------------
def _render_excel(doc_type, sheet, project, compliance_results, component_model):
    """Render an Excel workbook with openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Compliance Report"

    # Styles
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    title_font = Font(name="Calibri", bold=True, size=14, color="1E3A5F")
    info_label_font = Font(name="Calibri", bold=True, size=10)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    waived_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    pending_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

    title = _doc_title(doc_type, sheet, project)

    # Title
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1, value=title).font = title_font
    row += 1
    ws.cell(row=row, column=1, value="UnityESS Technical Compliance Portal").font = Font(
        name="Calibri", size=9, color="718096"
    )
    row += 2

    # Info section
    info_rows = []
    if project:
        info_rows.append(("Project", project.name))
        info_rows.append(("Client", project.client_name))
    if sheet:
        info_rows.append(("Sheet Number", sheet.sheet_number or "N/A"))
        stage_val = sheet.stage.value if hasattr(sheet.stage, "value") else str(sheet.stage)
        info_rows.append(("Stage", stage_val))
        info_rows.append(("Revision", str(sheet.revision)))
        info_rows.append(("Compliance Score", f"{sheet.compliance_score or 'N/A'}%"))
    if component_model:
        info_rows.append(("Component Model", component_model.model_name))
    info_rows.append(("Generated", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")))

    for label, value in info_rows:
        ws.cell(row=row, column=1, value=label).font = info_label_font
        ws.cell(row=row, column=2, value=value)
        row += 1
    row += 1

    # Results table
    if compliance_results:
        headers = ["Parameter ID", "OEM Value", "Status", "Notes"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
        row += 1

        for r in compliance_results:
            status_val = r.status.value if hasattr(r.status, "value") else str(r.status)
            if status_val == "pass" or (hasattr(r.status, "name") and r.status.name == "pass_"):
                status_display = "PASS"
                fill = pass_fill
            elif status_val == "fail":
                status_display = "FAIL"
                fill = fail_fill
            elif status_val == "waived":
                status_display = "WAIVED"
                fill = waived_fill
            else:
                status_display = "PENDING"
                fill = pending_fill

            param_id_short = str(r.standard_parameter_id)[:8] + "..."
            ws.cell(row=row, column=1, value=param_id_short).border = thin_border
            ws.cell(row=row, column=2, value=r.oem_value or "N/A").border = thin_border
            status_cell = ws.cell(row=row, column=3, value=status_display)
            status_cell.fill = fill
            status_cell.border = thin_border
            status_cell.alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=4, value=r.waive_reason or "").border = thin_border
            row += 1
    else:
        ws.cell(row=row, column=1, value="No compliance results available.")

    # Column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------
def _doc_title(doc_type: str, sheet, project) -> str:
    """Build a human-readable document title."""
    type_labels = {
        "compliance_sheet": "Compliance Sheet",
        "comparison_report": "Comparison Report",
        "customer_proposal": "Customer Proposal",
        "audit_export": "Audit Export",
        "full_project_pack": "Full Project Pack",
    }
    label = type_labels.get(doc_type, doc_type.replace("_", " ").title())
    parts = [label]
    if sheet and sheet.sheet_number:
        parts.append(f"— {sheet.sheet_number}")
    if project:
        parts.append(f"({project.name})")
    return " ".join(parts)


# ============================================================================
# TASK 2: run_ai_extraction
# ============================================================================
@celery_app.task(name="run_ai_extraction", bind=True, max_retries=2)
def run_ai_extraction(self, job_id: str):
    """
    Extract parameter values from a datasheet PDF using the Claude API.

    Workflow:
      1. Look up the ai_extraction_jobs row; set status='running'.
      2. Download the PDF from MinIO.
      3. Load all standard_parameters for the component type.
      4. Call Claude with a structured extraction prompt.
      5. Parse the response into parameter values with confidence scores.
      6. Upsert into parameter_values (skip manually verified values).
      7. Update the job row with counts and status='done'.
    """
    models = _import_models()
    AIExtractionJob = models["AIExtractionJob"]
    ComponentModel = models["ComponentModel"]
    ComponentType = models["ComponentType"]
    StandardParameter = models["StandardParameter"]
    ParameterValue = models["ParameterValue"]
    ExtractionSource = models["ExtractionSource"]

    session: Session = get_sync_session()
    try:
        # ------------------------------------------------------------------
        # 1. Load job and set status to running
        # ------------------------------------------------------------------
        job = session.get(AIExtractionJob, UUID(job_id))
        if job is None:
            raise ValueError(f"AI extraction job {job_id} not found")

        job.status = "running"
        session.commit()

        # Load the component model and its type
        comp_model = session.get(ComponentModel, job.component_model_id)
        if comp_model is None:
            raise ValueError(f"ComponentModel {job.component_model_id} not found")

        comp_type = session.get(ComponentType, comp_model.component_type_id)
        if comp_type is None:
            raise ValueError(f"ComponentType {comp_model.component_type_id} not found")

        # ------------------------------------------------------------------
        # 2. Download PDF from MinIO
        # ------------------------------------------------------------------
        s3 = _get_s3_client()
        bucket = settings.MINIO_BUCKET_DATASHEETS

        pdf_obj = s3.get_object(Bucket=bucket, Key=job.minio_key)
        pdf_bytes = pdf_obj["Body"].read()
        logger.info("Downloaded datasheet PDF (%d bytes) for job %s", len(pdf_bytes), job_id)

        # ------------------------------------------------------------------
        # 3. Load standard parameters for this component type
        # ------------------------------------------------------------------
        std_params = (
            session.execute(
                select(StandardParameter)
                .where(StandardParameter.component_type_id == comp_type.id)
                .order_by(StandardParameter.sort_order)
            )
            .scalars()
            .all()
        )

        if not std_params:
            raise ValueError(
                f"No standard parameters defined for component type '{comp_type.name}'"
            )

        # Build the parameter list for the prompt
        param_list_text = _build_parameter_prompt_list(std_params)

        # ------------------------------------------------------------------
        # 4. Call Claude API
        # ------------------------------------------------------------------
        import anthropic
        import base64

        client = anthropic.Anthropic(api_key=settings.ANTHROPIC_API_KEY)

        pdf_b64 = base64.standard_b64encode(pdf_bytes).decode("utf-8")

        system_prompt = (
            "You are a technical data extraction specialist for energy storage systems. "
            "You extract parameter values from component datasheets with high accuracy. "
            "You MUST respond ONLY with a valid JSON array — no markdown fences, no commentary."
        )

        user_prompt = f"""Analyze the attached PDF datasheet for a {comp_type.name} ({comp_type.category.value if hasattr(comp_type.category, 'value') else comp_type.category}) component.

Extract values for each of the following parameters. For each parameter, provide:
- "code": the parameter code (exactly as listed below)
- "value": the extracted value as a string (use the unit specified if applicable)
- "value_jsonb": for range types, provide {{"min": <number>, "max": <number>}}; otherwise null
- "confidence": a number between 0.0 and 1.0 indicating your confidence in the extraction
- "reasoning": brief explanation of where you found the value or why confidence is low

Parameters to extract:
{param_list_text}

Respond with a JSON array of objects. Example:
[
  {{"code": "NOM_VOLTAGE", "value": "3.2", "value_jsonb": null, "confidence": 0.95, "reasoning": "Found in electrical specifications table, row 'Nominal Voltage'"}},
  {{"code": "OPERATING_TEMP", "value": "-20 to 55", "value_jsonb": {{"min": -20, "max": 55}}, "confidence": 0.88, "reasoning": "Found in environmental section"}}
]

If a parameter value cannot be found in the document, include it with value=null and confidence=0.0.
Return ONLY the JSON array with no additional text."""

        response = client.messages.create(
            model=settings.CLAUDE_MODEL,
            max_tokens=4096,
            system=system_prompt,
            messages=[
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "document",
                            "source": {
                                "type": "base64",
                                "media_type": "application/pdf",
                                "data": pdf_b64,
                            },
                        },
                        {
                            "type": "text",
                            "text": user_prompt,
                        },
                    ],
                }
            ],
        )

        raw_text = response.content[0].text
        logger.info("Claude response received (%d chars) for job %s", len(raw_text), job_id)

        # ------------------------------------------------------------------
        # 5. Parse response
        # ------------------------------------------------------------------
        extracted_params = _parse_claude_response(raw_text)

        # Build a lookup: code -> StandardParameter
        param_by_code = {p.code: p for p in std_params}

        # ------------------------------------------------------------------
        # 6. Upsert parameter_values
        # ------------------------------------------------------------------
        count_extracted = 0
        count_auto_verified = 0
        count_flagged = 0
        count_manual = 0

        now = datetime.now(timezone.utc)

        for item in extracted_params:
            code = item.get("code")
            if not code or code not in param_by_code:
                continue

            std_param = param_by_code[code]
            value_raw = item.get("value")
            value_jsonb = item.get("value_jsonb")
            confidence = item.get("confidence", 0.0)

            if value_raw is None and confidence == 0.0:
                # Parameter not found in document — skip
                continue

            count_extracted += 1

            # Classify by confidence
            if confidence >= 0.90:
                count_auto_verified += 1
            elif confidence >= 0.80:
                count_flagged += 1
            else:
                count_manual += 1

            # Check if a manually verified value already exists — do not overwrite
            existing = session.execute(
                select(ParameterValue).where(
                    ParameterValue.component_model_id == comp_model.id,
                    ParameterValue.standard_parameter_id == std_param.id,
                )
            ).scalar_one_or_none()

            if existing and existing.is_verified:
                # Skip — manually verified values are sacrosanct
                logger.debug(
                    "Skipping verified param %s for model %s",
                    code, comp_model.id
                )
                continue

            if existing:
                # Update existing AI/unverified value
                existing.value_raw = str(value_raw) if value_raw is not None else None
                existing.value_jsonb = value_jsonb
                existing.confidence = confidence
                existing.source = ExtractionSource.ai
                existing.is_verified = False
                existing.extracted_at = now
                existing.updated_at = now
            else:
                # Insert new
                pv = ParameterValue(
                    component_model_id=comp_model.id,
                    standard_parameter_id=std_param.id,
                    value_raw=str(value_raw) if value_raw is not None else None,
                    value_jsonb=value_jsonb,
                    confidence=confidence,
                    source=ExtractionSource.ai,
                    is_verified=False,
                    extracted_at=now,
                )
                session.add(pv)

        # ------------------------------------------------------------------
        # 7. Update job to done
        # ------------------------------------------------------------------
        # Store raw response as JSON for audit (the column exists in schema
        # even if the ORM model may not map it — use raw SQL as fallback)
        try:
            job.raw_response = {"response_text": raw_text, "model": settings.CLAUDE_MODEL}
        except AttributeError:
            # raw_response not on ORM model — store via raw UPDATE
            from sqlalchemy import text as sa_text
            session.execute(
                sa_text(
                    "UPDATE ai_extraction_jobs SET raw_response = :resp WHERE id = :jid"
                ),
                {
                    "resp": json.dumps({"response_text": raw_text, "model": settings.CLAUDE_MODEL}),
                    "jid": str(job.id),
                },
            )

        job.status = "done"
        job.parameters_extracted = count_extracted
        job.parameters_auto_verified = count_auto_verified
        job.parameters_flagged = count_flagged
        job.parameters_manual = count_manual
        job.completed_at = now

        session.commit()
        logger.info(
            "AI extraction job %s completed: %d extracted, %d auto-verified, %d flagged, %d manual",
            job_id, count_extracted, count_auto_verified, count_flagged, count_manual,
        )

    except Exception as exc:
        session.rollback()
        # Record failure on the job row
        try:
            job = session.get(AIExtractionJob, UUID(job_id))
            if job:
                job.status = "failed"
                job.error_message = f"{type(exc).__name__}: {exc}"
                job.completed_at = datetime.now(timezone.utc)
                session.commit()
        except Exception:
            session.rollback()

        logger.exception("run_ai_extraction failed for job %s", job_id)
        raise self.retry(exc=exc, countdown=60 * (self.request.retries + 1))
    finally:
        session.close()


# ---------------------------------------------------------------------------
# AI extraction helpers
# ---------------------------------------------------------------------------
def _build_parameter_prompt_list(std_params) -> str:
    """Build a formatted list of parameters for the Claude prompt."""
    lines = []
    for p in std_params:
        data_type = p.data_type.value if hasattr(p.data_type, "value") else str(p.data_type)
        unit_str = f" (unit: {p.unit})" if p.unit else ""
        mandatory_str = " [MANDATORY]" if p.is_mandatory else ""
        section_str = f" [{p.section}]"

        constraints = ""
        if p.acceptance_min is not None and p.acceptance_max is not None:
            constraints = f" | acceptable range: {p.acceptance_min}–{p.acceptance_max}"
        elif p.acceptance_min is not None:
            constraints = f" | min: {p.acceptance_min}"
        elif p.acceptance_max is not None:
            constraints = f" | max: {p.acceptance_max}"
        if p.acceptance_enum:
            constraints += f" | allowed values: {', '.join(p.acceptance_enum)}"

        lines.append(
            f"- {p.code}: {p.display_name}{unit_str} (type: {data_type})"
            f"{section_str}{mandatory_str}{constraints}"
        )
    return "\n".join(lines)


def _parse_claude_response(raw_text: str) -> list[dict]:
    """Parse Claude's JSON response, handling markdown fences and minor issues."""
    text = raw_text.strip()

    # Strip markdown code fences if present
    if text.startswith("```"):
        # Remove opening fence (```json or ```)
        text = re.sub(r"^```(?:json)?\s*\n?", "", text)
        # Remove closing fence
        text = re.sub(r"\n?```\s*$", "", text)

    try:
        parsed = json.loads(text)
    except json.JSONDecodeError:
        # Try to find a JSON array in the text
        match = re.search(r"\[.*\]", text, re.DOTALL)
        if match:
            parsed = json.loads(match.group())
        else:
            logger.error("Failed to parse Claude response as JSON: %s...", text[:200])
            raise ValueError("Claude response is not valid JSON")

    if not isinstance(parsed, list):
        raise ValueError(f"Expected JSON array, got {type(parsed).__name__}")

    return parsed
